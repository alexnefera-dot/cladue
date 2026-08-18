#!/usr/bin/env python3
"""
Сборка базы брендов из исходного Excel в формат движка.

Вход: xlsx, где лист «Сводка» + по листу на бренд; у бренда колонки
      query | clicks (первая строка — заголовок).
Выход (в этой же папке):
      brands/<бренд>.json.gz  — [[запрос, клики], ...] по убыванию кликов
      brands-index.json       — индекс с маркерами авто-определения бренда

Запуск:
    pip install openpyxl
    python3 build-base.py /путь/к/brands_master.xlsx
"""
import sys, os, re, json, gzip, shutil
from collections import Counter

try:
    import openpyxl
except ImportError:
    sys.exit("Нужен openpyxl: pip install openpyxl")

GENERIC = set("""на не и в во с со из от до по о об за под над к ко у про для без при между через
    о а но да или либо что как когда пока если бы же ли вот тут там все весь это этот тот
    их его ее наш ваш мой твой the com ru www me top main online play win bet casino рус rus
    казино официальный сайт зеркало вход войти рабочее сегодня бет игровые автоматы играть
    онлайн промокод бонус регистрация кабинет скачать приложение версия контора казно бк
    личный мобильная мани money новое старое лучшее икс x""".split())

def toks(s):
    return [w for w in re.findall(r"[a-zа-я0-9]+", s.lower()) if len(w) >= 2]

def safe(n):
    return re.sub(r"[^a-z0-9_-]", "_", n.lower())

def main(src):
    here = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(here, "brands")
    if os.path.exists(out_dir):
        shutil.rmtree(out_dir)
    os.makedirs(out_dir)

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    index = []
    for name in wb.sheetnames:
        if name.strip().lower() == "сводка":
            continue
        ws = wb[name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or row[0] is None:
                continue
            q = str(row[0]).strip().lower()
            if not q:
                continue
            try:
                c = int(row[1]) if len(row) > 1 and row[1] is not None else 0
            except (TypeError, ValueError):
                c = 0
            rows.append([q, c])
        if not rows:
            continue
        rows.sort(key=lambda r: r[1], reverse=True)

        fn = safe(name)
        with gzip.open(f"{out_dir}/{fn}.json.gz", "wt", encoding="utf-8", compresslevel=9) as f:
            json.dump(rows, f, ensure_ascii=False)

        # маркеры авто-определения: различающие токены имени + кириллический вариант
        name_tokens = [t for t in re.split(r"[_\s]+", name.lower()) if t]
        cnt = Counter()
        for q, _ in rows[:120]:
            for w in toks(q):
                if w in GENERIC or w in name_tokens or len(w) < 3:
                    continue
                cnt[w] += 1
        cyr = [w for w, _ in cnt.most_common() if re.search(r"[а-я]", w) and len(w) >= 4]
        lat = [w for w, _ in cnt.most_common() if re.search(r"[a-z]", w) and len(w) >= 4]
        keys = [t for t in name_tokens if t not in GENERIC and (len(t) >= 3 or any(c.isdigit() for c in t))]
        if cyr:
            keys.append(cyr[0])
        if not keys and lat:
            keys.append(lat[0])

        index.append({
            "name": name, "file": fn,
            "total_clicks": sum(c for _, c in rows),
            "query_count": len(rows),
            "top_query": rows[0][0],
            "keys": keys,
        })

    index.sort(key=lambda b: b["total_clicks"], reverse=True)
    with open(os.path.join(here, "brands-index.json"), "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=0)

    total_q = sum(b["query_count"] for b in index)
    print(f"Готово: брендов={len(index)} запросов={total_q}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Использование: python3 build-base.py <brands_master.xlsx>")
    main(sys.argv[1])
