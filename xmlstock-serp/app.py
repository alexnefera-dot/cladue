#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
XMLStock SERP — локальный сборщик ТОП-выдачи Яндекса (и Google) через API xmlstock.com.

Запускает локальный веб-сервер с интерфейсом в браузере: вставляете пачку
запросов (200-300 шт.), получаете ТОП-N результатов по каждому ключу и
выгружаете в файл (XLSX / CSV).

Запуск:  python app.py
Открыть: http://127.0.0.1:5000  (откроется автоматически)
"""

import csv
import hmac
import io
import json
import os
import re
import shutil
import socket
import subprocess
import time
import uuid
import webbrowser
import xml.etree.ElementTree as ET
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from threading import Event, Lock, Thread
from urllib.parse import urljoin, urlparse

import requests
from flask import (
    Flask,
    Response,
    jsonify,
    render_template,
    request,
    send_file,
)

app = Flask(__name__)
# Перечитывать шаблон при изменении на диске и не давать браузеру кешировать
# страницу — чтобы после замены index.html хватило обновить вкладку (без рестарта).
app.config["TEMPLATES_AUTO_RELOAD"] = True

# Маркер сборки backend — показывается в футере. Если после обновления в футере
# старый маркер, значит сервер не перезапущен (app.py подхватывается только при рестарте).
APP_BUILD = "поз-и-url"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Анализ редиректов ходит по «мусорным» сайтам с битыми сертификатами —
# подавляем шум о непроверенном SSL (проверку можно включить в интерфейсе).
try:  # pragma: no cover
    import urllib3

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except Exception:
    pass

# Необязательная защита паролем — для случая, когда инструмент выставлен наружу
# (публичный поддомен). Включается ТОЛЬКО если задан AUTH_PASS в окружении; при
# локальном запуске на компьютере переменной нет — защита выключена, поведение
# не меняется. На сервере добавьте в systemd-юнит:
#   Environment=AUTH_USER=admin
#   Environment=AUTH_PASS=длинный_пароль
AUTH_USER = os.environ.get("AUTH_USER") or "admin"
AUTH_PASS = os.environ.get("AUTH_PASS") or ""


@app.before_request
def _require_auth():
    if not AUTH_PASS:
        return None
    a = request.authorization
    if (a and hmac.compare_digest(a.username or "", AUTH_USER)
            and hmac.compare_digest(a.password or "", AUTH_PASS)):
        return None
    return Response("Требуется авторизация.", 401,
                    {"WWW-Authenticate": 'Basic realm="xmlstock-serp"'})

# Снимок последнего «Сбора ТОП» — хранится на диске, переживает перезагрузку
# страницы и перезапуск сервера (пока не пересобрали новый).
LAST_COLLECT_FILE = os.path.join(OUTPUT_DIR, "last_collect.json")


def _save_last_collect(snap):
    try:
        with open(LAST_COLLECT_FILE, "w", encoding="utf-8") as f:
            json.dump(snap, f, ensure_ascii=False)
    except Exception:
        pass


def _load_last_collect():
    try:
        with open(LAST_COLLECT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

# Пресеты эндпоинтов xmlstock. У xmlstock РАЗНЫЕ адреса для живой выдачи (Live)
# и официального Яндекс.XML. Какой использовать — зависит от того, какие лимиты
# пополнены в аккаунте. Если адрес в кабинете другой — выберите «Свой URL».
#   Live  (живая выдача, ~10 результатов на запрос): /yandexlive/xml/
#   XML   (официальный Яндекс.XML, до 100 за запрос): /yandex/xml/
ENDPOINTS = {
    "yandex_live": "https://xmlstock.com/yandexlive/xml/",
    "yandex": "https://xmlstock.com/yandex/xml/",
    "google_live": "https://xmlstock.com/googlelive/xml/",
    "google": "https://xmlstock.com/google/xml/",
    # обратная совместимость со старыми сохранёнными значениями
    "yandex_xml": "https://xmlstock.com/yandex/xml/",
    "google_xml": "https://xmlstock.com/google/xml/",
}


def _is_live(url):
    """Эндпоинт живой выдачи (Live) — отдаёт максимум 10 результатов на запрос."""
    u = (url or "").lower()
    return "yandexlive" in u or "googlelive" in u or "/live/" in u

# Хранилище задач в памяти процесса.
JOBS = {}
JOBS_LOCK = Lock()

# Хранилище задач трекера позиций.
MONITORS = {}
MON_LOCK = Lock()
ACTIVE_MONITOR = {"id": None}

# Хранилище задач сбора страниц (оператор site:).
SITEJOBS = {}
SITE_LOCK = Lock()
ACTIVE_SITE = {"id": None}

# Хранилище задач анализа редиректов + указатель на последний «Сбор ТОП».
REDIRJOBS = {}
REDIR_LOCK = Lock()
ACTIVE_REDIR = {"id": None}
LAST_COLLECT = {"id": None}

# Хранилище задач проверки живости доменов.
LIVEJOBS = {}
LIVE_LOCK = Lock()
ACTIVE_LIVE = {"id": None}

# Хранилище задач мониторинга запусков (частые срезы + дата регистрации).
WATCHJOBS = {}
WATCH_LOCK = Lock()
ACTIVE_WATCH = {"id": None}


# --------------------------------------------------------------------------- #
#  Парсинг ответа                                                             #
# --------------------------------------------------------------------------- #
def domain_from_url(url):
    try:
        netloc = urlparse(url).netloc.lower()
        return netloc[4:] if netloc.startswith("www.") else netloc
    except Exception:
        return ""


def normalize_domain(value):
    """Приводит «www.Site.ru/path» или «https://site.ru» к «site.ru»."""
    s = (value or "").strip().lower()
    if not s:
        return ""
    if "://" in s:
        s = urlparse(s).netloc or s.split("://", 1)[1]
    s = s.split("/")[0].strip()
    if s.startswith("www."):
        s = s[4:]
    return s


def _domain_matches(found, target):
    """Совпадение домена в любую сторону: точное, либо один является
    поддоменом другого. Так сайт ловится и если ввели корень (k4r.team —
    поймает spacewin.k4r.team), и если ввели поддомен (spacewin.k4r.team —
    поймает и корень k4r.team). Путь/раздел и www роли не играют."""
    if not found or not target:
        return False
    found = found.lower()
    if found.startswith("www."):
        found = found[4:]
    return (
        found == target
        or found.endswith("." + target)
        or target.endswith("." + found)
    )


def find_position(results, target):
    """Позиция целевого домена в выдаче (1-based) или None, если не найден."""
    for i, r in enumerate(results, 1):
        dom = r.get("domain") or domain_from_url(r.get("url", ""))
        if _domain_matches(dom, target):
            return i
    return None


def find_position_url(results, target):
    """Позиция И полный URL ранжирующейся страницы домена. (None, None) если нет."""
    for i, r in enumerate(results, 1):
        dom = r.get("domain") or domain_from_url(r.get("url", ""))
        if _domain_matches(dom, target):
            return i, (r.get("url") or "")
    return None, None


def _is_retryable(error_text):
    """Временные ошибки, которые имеет смысл повторить."""
    if not error_text:
        return False
    t = error_text.lower()
    return any(
        s in t
        for s in ("limit", "лимит", "rps", "temporar", "временно", "503", "502", "504", "busy")
    )


def parse_xml(text):
    """Разбор ответа в формате Яндекс XML. Возвращает (results, error)."""
    try:
        root = ET.fromstring(text)
    except ET.ParseError as e:
        return [], f"ошибка XML: {e}"

    err = root.find(".//response/error")
    if err is None:
        err = root.find(".//error")
    if err is not None:
        code = (err.get("code") or "").strip()
        msg = (err.text or "").strip()
        # 15 — «по запросу ничего не нашлось», это не ошибка, просто пусто.
        if code == "15":
            return [], None
        return [], f"API {code}: {msg}".strip()

    results = []
    for doc in root.findall(".//doc"):
        url = (doc.findtext("url") or "").strip()
        if not url:
            continue
        title_el = doc.find("title")
        title = "".join(title_el.itertext()).strip() if title_el is not None else ""
        domain = (doc.findtext("domain") or "").strip() or domain_from_url(url)
        results.append({"url": url, "title": title, "domain": domain})
    return results, None


def parse_json(data):
    """Запасной разбор JSON: рекурсивно собираем объекты с полем url."""
    found = []

    def walk(node):
        if isinstance(node, dict):
            if isinstance(node.get("url"), str):
                found.append(node)
            for v in node.values():
                walk(v)
        elif isinstance(node, list):
            for v in node:
                walk(v)

    walk(data)
    results = []
    for o in found:
        url = (o.get("url") or "").strip()
        if not url:
            continue
        title = o.get("title") or o.get("name") or ""
        if isinstance(title, dict):
            title = title.get("text", "")
        domain = (o.get("domain") or "").strip() or domain_from_url(url)
        results.append({"url": url, "title": str(title).strip(), "domain": domain})
    if not results:
        return [], "не удалось разобрать JSON-ответ"
    return results, None


def parse_response(text):
    text = (text or "").strip()
    if not text:
        return [], "пустой ответ"
    if text[0] == "<":
        return parse_xml(text)
    if text[0] in "{[":
        try:
            return parse_json(json.loads(text))
        except Exception:
            pass
    return [], "нераспознанный ответ: " + text[:200].replace("\n", " ")


# --------------------------------------------------------------------------- #
#  Запрос к API                                                               #
# --------------------------------------------------------------------------- #
def _fetch_page(cfg, query, page, page_size, cancel):
    """Один HTTP-запрос (одна страница) с ретраями. Возвращает (results, error)."""
    groupby = f"attr=d.mode=deep.groups-on-page={page_size}.docs-in-group=1"
    params = {
        "user": cfg["user"],
        "key": cfg["key"],
        "query": query,
        "groupby": groupby,
        "page": page,
    }
    if cfg.get("lr"):
        params["lr"] = cfg["lr"]
    if cfg.get("domain"):
        params["domain"] = cfg["domain"]
    if cfg.get("device"):
        params["device"] = cfg["device"]

    last_err = None
    for attempt in range(cfg["retries"] + 1):
        if cancel.is_set():
            return [], "отменено"
        try:
            r = requests.get(cfg["endpoint"], params=params, timeout=cfg["timeout"])
        except requests.RequestException as e:
            last_err = f"сеть: {e}"
            time.sleep(min(2 ** attempt, 8))
            continue

        if r.status_code >= 500:
            last_err = f"HTTP {r.status_code}"
            time.sleep(min(2 ** attempt, 8))
            continue
        if r.status_code != 200:
            return [], f"HTTP {r.status_code}: {r.text[:200]}"

        results, err = parse_response(r.text)
        if err and _is_retryable(err) and attempt < cfg["retries"]:
            last_err = err
            time.sleep(min(2 ** attempt, 8))
            continue
        return results, err

    return [], last_err or "не удалось выполнить запрос"


def fetch_one(cfg, query, cancel):
    """ТОП-N результатов по запросу. Живая выдача (Live) xmlstock отдаёт максимум
    10 результатов на запрос, поэтому при большей глубине листаем страницы.
    XML отдаёт до 100 за один запрос."""
    if cancel.is_set():
        return [], "отменено"

    top_n = max(1, cfg["top_n"])
    page_size = 10 if cfg.get("live") else min(top_n, 100)
    pages = min((top_n + page_size - 1) // page_size, 10)  # предохранитель

    collected, last_err = [], None
    for p in range(pages):
        if cancel.is_set():
            break
        if cfg["delay"] > 0:
            time.sleep(cfg["delay"])
        results, err = _fetch_page(cfg, query, p, page_size, cancel)
        if err:
            if not collected:
                return [], err
            last_err = err
            break
        if not results:
            break
        collected.extend(results)
        if len(results) < page_size or len(collected) >= top_n:
            break

    return collected[:top_n], (None if collected else last_err)


# --------------------------------------------------------------------------- #
#  Сбор страниц сайта из индекса (оператор site:)                             #
# --------------------------------------------------------------------------- #
def clean_host(value):
    """Хост для оператора site: — убираем схему/путь, но СОХРАНЯЕМ www/поддомен."""
    s = (value or "").strip().lower()
    if not s:
        return ""
    if "://" in s:
        s = s.split("://", 1)[1]
    return s.split("/")[0].strip()


def parse_site(text):
    """Разбор ответа для site:. Возвращает (docs[{url,title}], found:int|None, error)."""
    text = (text or "").strip()
    if not text:
        return [], None, "пустой ответ"
    if text[0] != "<":
        docs, err = parse_response(text)
        return docs, None, err
    try:
        root = ET.fromstring(text)
    except ET.ParseError as e:
        return [], None, f"ошибка XML: {e}"
    err = root.find(".//response/error") or root.find(".//error")
    if err is not None:
        code = (err.get("code") or "").strip()
        if code == "15":  # ничего не нашлось
            return [], 0, None
        return [], None, f"API {code}: {(err.text or '').strip()}"
    found = None
    for f in root.findall(".//found"):
        if f.get("priority") == "phrase" and (f.text or "").strip().isdigit():
            found = int(f.text.strip())
            break
    if found is None:
        f = root.find(".//found")
        if f is not None and (f.text or "").strip().isdigit():
            found = int(f.text.strip())
    docs = []
    for doc in root.findall(".//doc"):
        url = (doc.findtext("url") or "").strip()
        if not url:
            continue
        te = doc.find("title")
        title = "".join(te.itertext()).strip() if te is not None else ""
        docs.append({"url": url, "title": title})
    return docs, found, None


def _fetch_site_page(cfg, query, page, page_size, cancel):
    # Плоская выдача (mode=flat): для site: нельзя схлопывать по домену.
    groupby = f'attr="".mode=flat.groups-on-page={page_size}.docs-in-group=1'
    params = {"user": cfg["user"], "key": cfg["key"], "query": query,
              "groupby": groupby, "page": page}
    if cfg.get("lr"):
        params["lr"] = cfg["lr"]
    if cfg.get("device"):
        params["device"] = cfg["device"]
    last_err = None
    for attempt in range(cfg["retries"] + 1):
        if cancel.is_set():
            return [], None, "отменено"
        try:
            r = requests.get(cfg["endpoint"], params=params, timeout=cfg["timeout"])
        except requests.RequestException as e:
            last_err = f"сеть: {e}"; time.sleep(min(2 ** attempt, 8)); continue
        if r.status_code >= 500:
            last_err = f"HTTP {r.status_code}"; time.sleep(min(2 ** attempt, 8)); continue
        if r.status_code != 200:
            return [], None, f"HTTP {r.status_code}: {r.text[:200]}"
        docs, found, err = parse_site(r.text)
        if err and _is_retryable(err) and attempt < cfg["retries"]:
            last_err = err; time.sleep(min(2 ** attempt, 8)); continue
        return docs, found, err
    return [], None, last_err or "не удалось выполнить запрос"


def _fetch_site_query(cfg, query, cap, cancel):
    """Пагинация по ОДНОМУ запросу. Возвращает {urls, found, error}."""
    page_size = 10 if cfg.get("live") else 100
    max_pages = min((cap + page_size - 1) // page_size, 100)
    urls, seen, found_total, err_out = [], set(), None, None
    for p in range(max_pages):
        if cancel.is_set():
            break
        if cfg["delay"] > 0:
            time.sleep(cfg["delay"])
        docs, found, err = _fetch_site_page(cfg, query, p, page_size, cancel)
        if err:
            if not urls:
                err_out = err
            break
        if found_total is None:
            found_total = found
        if not docs:
            break
        new = 0
        for d in docs:
            if d["url"] not in seen:
                seen.add(d["url"])
                urls.append(d)
                new += 1
        if len(docs) < page_size or len(urls) >= cap or new == 0:
            break
    return {"urls": urls[:cap], "found": found_total, "error": err_out}


def _segments_at(docs, index, limit, min_count=2):
    """Сегменты пути на позиции index (0=раздел, 1=подраздел) по частоте.
    Берём только те, что встречаются >= min_count раз (это разделы, а не листовые
    ID-страницы вроде /game/12345) — чтобы не дробить в пустоту."""
    freq = {}
    for d in docs:
        try:
            parts = [p for p in urlparse(d["url"]).path.split("/") if p]
        except Exception:
            continue
        if len(parts) > index:
            seg = parts[index].strip()
            if seg:
                freq[seg] = freq.get(seg, 0) + 1
    segs = [(s, c) for s, c in freq.items() if c >= min_count]
    return [s for s, _ in sorted(segs, key=lambda x: -x[1])[:limit]]


def _collect_site(cfg, op, host, path_segs, cap, cancel, urls, seen, depth, budget):
    """Рекурсивно собирает site:домен[/раздел[/подраздел]], дробя при потолке."""
    if cancel.is_set() or len(urls) >= cap or budget[0] <= 0:
        return None, None
    prefix = f"{op}{host}"
    if path_segs:
        prefix += "/" + "/".join(path_segs)
    budget[0] -= 1
    res = _fetch_site_query(cfg, prefix, min(cap, 1000), cancel)
    got = res["urls"]
    for d in got:
        if d["url"] not in seen:
            seen.add(d["url"])
            urls.append(d)
            if len(urls) >= cap:
                break
    found, err = res["found"], res["error"]
    hit_cap = len(got) >= 100                       # запрос отдал полную порцию
    more = (found is None) or (found > len(got))    # Яндекс говорит, что есть ещё
    if (cfg.get("deep") and not err and depth < cfg.get("deep_levels", 2)
            and hit_cap and more and len(urls) < cap and budget[0] > 0
            and not cancel.is_set()):
        for seg in _segments_at(got, len(path_segs), 30):
            if cancel.is_set() or len(urls) >= cap or budget[0] <= 0:
                break
            _collect_site(cfg, op, host, path_segs + [seg], cap, cancel, urls, seen, depth + 1, budget)
    return found, err


def _reverse_host(host):
    parts = [p for p in (host or "").split(".") if p]
    return ".".join(reversed(parts))


def _operator_query(opname, host):
    if opname == "host":
        return f"host:{host}"
    if opname == "url":
        return f"url:{host}*"
    if opname == "rhost":
        return f"rhost:{_reverse_host(host)}.*"
    if opname == "plain":
        return host
    return f"site:{host}"


def _collect_for_operator(cfg, opname, host, cap, cancel, budget):
    """URL для одного оператора. site/host — с рекурсивным дроблением по пути,
    остальные (url/rhost/домен) — одиночным запросом."""
    urls, seen = [], set()
    if opname in ("site", "host"):
        found, err = _collect_site(cfg, opname + ":", host, [], cap, cancel, urls, seen, 0, budget)
    else:
        found, err = None, None
        if budget[0] > 0 and not cancel.is_set():
            budget[0] -= 1
            res = _fetch_site_query(cfg, _operator_query(opname, host), min(cap, 1000), cancel)
            for d in res["urls"]:
                if d["url"] not in seen:
                    seen.add(d["url"])
                    urls.append(d)
            found, err = res["found"], res["error"]
    return urls, found, err


def fetch_site_pages(cfg, domain, cap, cancel):
    """Собирает URL страниц домена из индекса. «Все операторы» гоняет несколько
    конструкций (site/host/url/rhost/домен) и объединяет уникальные срезы индекса.
    Затем перебирает найденные ПОДДОМЕНЫ по отдельности (site:поддомен) — страницы
    часто размазаны по ним, а общий запрос упирается в потолок ~200."""
    host = clean_host(domain)
    target = registrable_domain(host)
    base_op = "host" if cfg.get("operator") == "host:" else "site"
    multi = cfg.get("multi_op")
    sub_limit = cfg.get("deep_subdomains", 150)
    ops = ["site", "host", "url", "rhost", "plain"] if multi else [base_op]
    budget = [cfg.get("deep_budget", (sub_limit * 2 + 60) if multi else 60)]

    union_seen, union_urls, per_op = set(), [], {}

    def merge(local):
        cnt = 0
        for d in local:
            if registrable_domain(url_host(d["url"])) != target:  # чужие URL — мимо
                continue
            cnt += 1
            if d["url"] not in union_seen:
                union_seen.add(d["url"])
                union_urls.append(d)
                if len(union_urls) >= cap:
                    break
        return cnt

    base_found, base_err = None, None
    for opname in ops:
        if cancel.is_set() or len(union_urls) >= cap or budget[0] <= 0:
            break
        local, found, err = _collect_for_operator(cfg, opname, host, cap, cancel, budget)
        per_op[opname] = merge(local)
        if opname == base_op:
            base_found, base_err = found, err

    # Перебор поддоменов: страницы часто на поддоменах (casino.домен и т.п.),
    # а site:/rhost: упираются в общий потолок ~200. Добираем каждый по отдельности.
    if (multi or cfg.get("deep")) and not cancel.is_set() and len(union_urls) < cap and budget[0] > 0:
        subs, subseen = [], set()
        for d in list(union_urls):
            h = url_host(d["url"])
            if (h and h != host and h not in subseen
                    and registrable_domain(h) == target and is_subdomain(h)):
                subseen.add(h)
                subs.append(h)
        added = 0
        for sub in subs[:sub_limit]:
            if cancel.is_set() or len(union_urls) >= cap or budget[0] <= 0:
                break
            local, _, _ = _collect_for_operator(cfg, "site", sub, cap, cancel, budget)
            added += merge(local)
        if subs:
            per_op["поддомены"] = added

    return {"domain": domain, "host": host, "urls": union_urls[:cap],
            "found": base_found, "error": base_err,
            "collected": min(len(union_urls), cap), "per_op": per_op}


def run_site(job_id, cfg, domains):
    job = SITEJOBS[job_id]
    try:
        with ThreadPoolExecutor(max_workers=cfg["workers"]) as ex:
            futs = {ex.submit(fetch_site_pages, cfg, d, cfg["max_urls"], job["cancel"]): d
                    for d in domains}
            for fut in as_completed(futs):
                d = futs[fut]
                try:
                    res = fut.result()
                except Exception as e:  # noqa: BLE001
                    res = {"domain": d, "host": d, "urls": [], "found": None,
                           "error": f"исключение: {e}", "collected": 0}
                with job["lock"]:
                    job["done"] += 1
                    job["results"][d] = res
                    if res.get("error"):
                        job["errors"].append({"domain": d, "error": res["error"]})
    except Exception as e:  # noqa: BLE001
        with job["lock"]:
            job["status"] = "error"
            job["fatal"] = str(e)
        return
    with job["lock"]:
        job["status"] = "cancelled" if job["cancel"].is_set() else "done"
        job["finished_at"] = time.time()


# --------------------------------------------------------------------------- #
#  Анализ редиректов (куда ведут сайты из выдачи)                             #
# --------------------------------------------------------------------------- #
_TWO_LEVEL_SLD = {"co", "com", "net", "org", "gov", "edu", "ac", "or", "go", "pp"}


def registrable_domain(host):
    """Корневой (регистрируемый) домен без поддоменов. Эвристика последних 2-3 меток."""
    host = (host or "").lower().strip().strip(".")
    if host.startswith("www."):
        host = host[4:]
    parts = [p for p in host.split(".") if p]
    if len(parts) <= 2:
        return ".".join(parts)
    if parts[-2] in _TWO_LEVEL_SLD:
        return ".".join(parts[-3:])
    return ".".join(parts[-2:])


def is_subdomain(host):
    """True, если это поддомен (а не корневой домен и не www.корень)."""
    h = (host or "").lower().strip().strip(".")
    if h.startswith("www."):
        h = h[4:]
    return bool(h) and h != registrable_domain(host)


def url_host(url):
    try:
        netloc = urlparse(url if "://" in url else "http://" + url).netloc.lower()
        return netloc.split("@")[-1].split(":")[0]
    except Exception:
        return ""


_META_REFRESH_RE = re.compile(
    r'<meta[^>]+http-equiv=["\']?refresh["\']?[^>]*content=["\']?\s*\d+\s*;\s*url=([^"\'>\s]+)',
    re.I,
)
_JS_LOC_RE = re.compile(
    r'(?:(?:window|document|top|self|parent)\s*\.\s*)?location\s*'
    r'(?:\.\s*(?:href|assign|replace)\s*)?\s*(?:=|\()\s*["\']([^"\']+)["\']',
    re.I,
)


def find_client_redirect(html, base):
    """Мета-refresh или простой JS-редирект (window.location=...) на другой хост."""
    if not html:
        return None
    m = _META_REFRESH_RE.search(html)
    if m:
        return urljoin(base, m.group(1).strip())
    m = _JS_LOC_RE.search(html[:40000])
    if m:
        tgt = m.group(1).strip()
        if tgt and not tgt.startswith("#") and "javascript:" not in tgt.lower():
            return urljoin(base, tgt)
    return None


def resolve_redirect(url, cfg, cancel):
    """Идёт по редиректам (HTTP 3xx + meta/JS) и возвращает финальный URL."""
    if cancel.is_set():
        return {"final_url": url, "status": None, "error": "отменено"}
    headers = {"User-Agent": cfg["ua"], "Accept": "text/html,application/xhtml+xml,*/*",
               "Accept-Language": "ru,en;q=0.8"}
    if cfg.get("referer"):
        headers["Referer"] = cfg["referer"]
    current = url if "://" in url else "http://" + url
    status, err = None, None
    try:
        for _ in range(4):  # до 4 клиентских хопов (meta/JS)
            if cancel.is_set():
                break
            r = requests.get(current, headers=headers, timeout=cfg["timeout"],
                             allow_redirects=True, verify=cfg["verify"])
            status = r.status_code
            final = r.url
            ctype = (r.headers.get("Content-Type") or "").lower()
            nxt = find_client_redirect(r.text, final) if "html" in ctype else None
            if nxt and url_host(nxt) and url_host(nxt) != url_host(final):
                current = nxt
                continue
            current = final
            break
    except requests.RequestException as e:
        err = str(e)[:200]
    return {"final_url": current, "status": status, "error": err}


def _redir_row(it, res):
    final_host = url_host(res["final_url"])
    final_dom = registrable_domain(final_host)
    src_dom = registrable_domain(it["host"])
    return {
        "source_url": it["url"], "source_host": it["host"],
        "final_url": res["final_url"], "final_host": final_host,
        "final_domain": final_dom, "redirected": bool(final_dom and final_dom != src_dom),
        "status": res["status"], "error": res["error"],
    }


def _try_click(page):
    """Best-effort клик по заметной кнопке/ссылке (для дорвеев с кнопкой-входом)."""
    for w in ["играть", "войти", "продолж", "перейти", "начать", "получить",
              "бонус", "play", "enter", "casino", "claim", "get", "continue"]:
        try:
            loc = page.get_by_text(re.compile(w, re.I))
            if loc.count() > 0:
                loc.first.click(timeout=1500)
                return True
        except Exception:
            continue
    try:
        page.locator("a, button").first.click(timeout=1500)
        return True
    except Exception:
        return False


def _browser_resolve(ctx, url, cfg):
    """Открывает страницу в реальном браузере и ждёт смены домена (JS/таймер/клик)."""
    page = ctx.new_page()
    src = registrable_domain(url_host(url))
    try:
        try:
            page.goto(url if "://" in url else "http://" + url,
                      wait_until="commit", timeout=cfg["timeout"] * 1000)
        except Exception:
            pass  # редирект-цепочка может «рвать» goto — дальше следим по page.url
        waited, step = 0, 300
        while waited < cfg["wait_ms"]:
            cur = registrable_domain(url_host(page.url))
            if cur and cur != src:
                break
            page.wait_for_timeout(step)
            waited += step
        final = page.url
        if cfg.get("click") and registrable_domain(url_host(final)) == src:
            if _try_click(page):
                page.wait_for_timeout(min(cfg["wait_ms"], 4000))
                final = page.url
        return {"final_url": final, "status": None, "error": None}
    except Exception as e:  # noqa: BLE001
        return {"final_url": url, "status": None, "error": str(e)[:200]}
    finally:
        try:
            page.close()
        except Exception:
            pass


def _run_redirects_http(job, cfg, items):
    with ThreadPoolExecutor(max_workers=cfg["workers"]) as ex:
        futs = {ex.submit(resolve_redirect, it["url"], cfg, job["cancel"]): it for it in items}
        for fut in as_completed(futs):
            it = futs[fut]
            try:
                res = fut.result()
            except Exception as e:  # noqa: BLE001
                res = {"final_url": it["url"], "status": None, "error": str(e)[:200]}
            with job["lock"]:
                job["done"] += 1
                job["rows"].append(_redir_row(it, res))


def _run_redirects_browser(job, cfg, items):
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        with job["lock"]:
            job["status"] = "error"
            job["fatal"] = ("Браузерный режим требует Playwright. В папке проекта выполните: "
                            "pip install playwright && playwright install chromium")
        return
    launch = {"headless": True}
    exe = os.environ.get("PW_CHROMIUM_PATH")
    if exe:
        launch["executable_path"] = exe
    mobile = ("iPhone" in cfg["ua"]) or ("Mobile" in cfg["ua"])
    with sync_playwright() as p:
        try:
            browser = p.chromium.launch(**launch)
        except Exception as e:  # noqa: BLE001
            with job["lock"]:
                job["status"] = "error"
                job["fatal"] = ("Не удалось запустить Chromium: " + str(e)[:120] +
                                " — выполните: playwright install chromium")
            return
        ctx = browser.new_context(
            user_agent=cfg["ua"], ignore_https_errors=not cfg["verify"], locale="ru-RU",
            viewport={"width": 390, "height": 844} if mobile else {"width": 1366, "height": 768},
        )
        if cfg.get("referer"):
            ctx.set_extra_http_headers({"Referer": cfg["referer"]})
        for it in items:
            if job["cancel"].is_set():
                break
            res = _browser_resolve(ctx, it["url"], cfg)
            with job["lock"]:
                job["done"] += 1
                job["rows"].append(_redir_row(it, res))
        try:
            browser.close()
        except Exception:
            pass


def run_redirects(job_id, cfg, urls):
    job = REDIRJOBS[job_id]
    # Дедуп по ключу группировки (корневой домен или хост с поддоменом).
    key_root = cfg["group_by"] == "root"
    only_sub = cfg.get("only_subdomains", True)
    seen, items = set(), []
    for u in urls:
        h = url_host(u)
        if not h:
            continue
        if only_sub and not is_subdomain(h):  # только поддомены
            continue
        k = registrable_domain(h) if key_root else h
        if k and k not in seen:  # дедуп по корневому домену (или по хосту)
            seen.add(k)
            items.append({"url": u, "host": h})
    with job["lock"]:
        job["total"] = len(items)

    try:
        if cfg.get("engine_mode") == "browser":
            _run_redirects_browser(job, cfg, items)
        else:
            _run_redirects_http(job, cfg, items)
    except Exception as e:  # noqa: BLE001
        with job["lock"]:
            job["status"] = "error"
            job["fatal"] = str(e)[:300]
        return
    with job["lock"]:
        if job["status"] == "running":
            job["status"] = "cancelled" if job["cancel"].is_set() else "done"
            job["finished_at"] = time.time()


def redir_summary(rows):
    total = len(rows)
    counts, no_redir, errs = {}, 0, 0
    for r in rows:
        if r.get("error"):
            errs += 1
        elif r["redirected"] and r["final_domain"]:
            counts[r["final_domain"]] = counts.get(r["final_domain"], 0) + 1
        else:
            no_redir += 1

    def pct(n):
        return round(n / total * 100, 1) if total else 0

    targets = [{"domain": d, "count": c, "percent": pct(c)}
               for d, c in sorted(counts.items(), key=lambda x: -x[1])]
    return {"total": total, "targets": targets,
            "no_redirect": no_redir, "no_redirect_percent": pct(no_redir),
            "errors": errs, "errors_percent": pct(errs)}


# --------------------------------------------------------------------------- #
#  Проверка живости доменов (жив / мёртв)                                     #
# --------------------------------------------------------------------------- #
def _live_note(code, host, final_url):
    fh = url_host(final_url)
    if fh and registrable_domain(fh) != registrable_domain(host):
        return "редирект на " + registrable_domain(fh)
    if code and code >= 400:
        return f"код {code}"
    return "ок"


def check_alive(domain, cfg, cancel):
    """DNS + HTTP-проверка. Возвращает статус: жив / мёртв (нет DNS / не отвечает)."""
    host = url_host(domain) or clean_host(domain)
    base = {"domain": domain, "host": host}
    if cancel.is_set():
        return {**base, "alive": False, "status": None, "final_url": "", "note": "отменено", "error": None}
    if not host:
        return {**base, "alive": False, "status": None, "final_url": "", "note": "плохой домен", "error": None}
    try:
        socket.gethostbyname(host)
    except Exception as e:  # noqa: BLE001
        return {**base, "alive": False, "status": None, "final_url": "",
                "note": "нет DNS", "error": str(e)[:120]}
    headers = {"User-Agent": cfg["ua"], "Accept": "text/html,*/*"}
    last = None
    for scheme in ("https://", "http://"):
        if cancel.is_set():
            break
        try:
            r = requests.get(scheme + host, headers=headers, timeout=cfg["timeout"],
                             allow_redirects=True, verify=False, stream=True)
            code, final = r.status_code, r.url
            try:
                r.close()
            except Exception:
                pass
            return {**base, "alive": True, "status": code, "final_url": final,
                    "note": _live_note(code, host, final), "error": None}
        except requests.RequestException as e:
            last = str(e)[:120]
    return {**base, "alive": False, "status": None, "final_url": "",
            "note": "не отвечает", "error": last}


def run_live(job_id, cfg, domains):
    job = LIVEJOBS[job_id]
    try:
        with ThreadPoolExecutor(max_workers=cfg["workers"]) as ex:
            futs = {ex.submit(check_alive, d, cfg, job["cancel"]): d for d in domains}
            for fut in as_completed(futs):
                try:
                    res = fut.result()
                except Exception as e:  # noqa: BLE001
                    d = futs[fut]
                    res = {"domain": d, "host": d, "alive": False, "status": None,
                           "final_url": "", "note": "ошибка", "error": str(e)[:120]}
                with job["lock"]:
                    job["done"] += 1
                    job["rows"].append(res)
    except Exception as e:  # noqa: BLE001
        with job["lock"]:
            job["status"] = "error"
            job["fatal"] = str(e)[:200]
        return
    with job["lock"]:
        job["status"] = "cancelled" if job["cancel"].is_set() else "done"
        job["finished_at"] = time.time()


# --------------------------------------------------------------------------- #
#  Фоновая обработка пачки                                                    #
# --------------------------------------------------------------------------- #
def run_job(job_id, cfg, queries):
    job = JOBS[job_id]
    results_by_query = {}

    try:
        with ThreadPoolExecutor(max_workers=cfg["workers"]) as ex:
            futures = {}
            for q in queries:
                if job["cancel"].is_set():
                    break
                futures[ex.submit(fetch_one, cfg, q, job["cancel"])] = q

            for fut in as_completed(futures):
                q = futures[fut]
                try:
                    res, err = fut.result()
                except Exception as e:  # noqa: BLE001
                    res, err = [], f"исключение: {e}"
                with job["lock"]:
                    job["done"] += 1
                    results_by_query[q] = (res, err)
                    if err:
                        job["errors"].append({"query": q, "error": err})
                    job["recent"] = (
                        job["recent"] + [{"query": q, "count": len(res), "error": err}]
                    )[-15:]
    except Exception as e:  # noqa: BLE001
        with job["lock"]:
            job["status"] = "error"
            job["fatal"] = str(e)
        return

    # Сборка файлов в исходном порядке запросов.
    long_rows = []
    wide_rows = []
    for q in queries:
        res, err = results_by_query.get(q, ([], "не обработано"))
        if res:
            for i, r in enumerate(res, 1):
                long_rows.append([q, i, r["url"], r["title"], r["domain"]])
            urls = [r["url"] for r in res]
        else:
            note = f"[{err}]" if err else "[нет результатов]"
            long_rows.append([q, "", "", note, ""])
            urls = []
        urls = (urls + [""] * cfg["top_n"])[: cfg["top_n"]]
        wide_rows.append([q] + urls)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = os.path.join(OUTPUT_DIR, f"serp_{stamp}.csv")
    xlsx_path = os.path.join(OUTPUT_DIR, f"serp_{stamp}.xlsx")

    _write_csv(csv_path, long_rows)
    xlsx_ok = _write_xlsx(xlsx_path, long_rows, wide_rows, cfg["top_n"])

    result_urls = [r[2] for r in long_rows if r[2]]
    with job["lock"]:
        job["csv_path"] = csv_path
        job["xlsx_path"] = xlsx_path if xlsx_ok else None
        job["rows"] = len(long_rows)
        job["result_urls"] = result_urls
        job["status"] = "cancelled" if job["cancel"].is_set() else "done"
        job["finished_at"] = time.time()
    _save_last_collect({
        "job_id": job_id, "at": time.time(),
        "total": job["total"], "rows": len(long_rows),
        "urls": result_urls,
        "csv": csv_path, "xlsx": xlsx_path if xlsx_ok else None,
    })


def _write_csv(path, long_rows):
    # utf-8-sig + разделитель «;» — корректно открывается двойным кликом в Excel (RU).
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["query", "position", "url", "title", "domain"])
        w.writerows(long_rows)


def _write_xlsx(path, long_rows, wide_rows, top_n):
    try:
        from openpyxl import Workbook
    except Exception:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(["query", "position", "url", "title", "domain"])
    for row in long_rows:
        ws.append(row)

    ws2 = wb.create_sheet("wide")
    ws2.append(["query"] + [f"url_{i}" for i in range(1, top_n + 1)])
    for row in wide_rows:
        ws2.append(row)

    try:
        wb.save(path)
        return True
    except Exception:
        return False


# --------------------------------------------------------------------------- #
#  Трекер позиций: прогон по раундам                                          #
# --------------------------------------------------------------------------- #
def run_monitor(job_id, cfg_base):
    job = MONITORS[job_id]
    sources = job["sources"]
    keywords = job["keywords"]
    pairs_by_kw = job["pairs_by_kw"]
    # Отдельный cfg на каждый источник (Live/XML отличаются endpoint и режимом).
    scfg = {
        s["name"]: {**cfg_base, "endpoint": s["endpoint"], "live": s["live"]}
        for s in sources
    }
    try:
        for rnd in range(1, job["rounds_total"] + 1):
            if job["cancel"].is_set():
                break
            round_pos = {s["name"]: {} for s in sources}
            tasks = [(s["name"], kw) for s in sources for kw in keywords]
            with ThreadPoolExecutor(max_workers=cfg_base["workers"]) as ex:
                futs = {
                    ex.submit(fetch_one, scfg[sname], kw, job["cancel"]): (sname, kw)
                    for (sname, kw) in tasks
                }
                for fut in as_completed(futs):
                    sname, kw = futs[fut]
                    try:
                        results, err = fut.result()
                    except Exception as e:  # noqa: BLE001
                        results, err = [], f"исключение: {e}"
                    if err:
                        with job["lock"]:
                            job["errors"].append(
                                {"round": rnd, "source": sname, "query": kw, "error": err}
                            )
                    doms = pairs_by_kw.get(kw, [])
                    round_pos[sname][kw] = {dom: find_position(results, dom) for dom in doms}

            with job["lock"]:
                job["snapshots"].append(
                    {"round": rnd, "at": time.time(), "positions": round_pos}
                )
                job["rounds_done"] = rnd

            # Пауза до следующего снятия (прерываемая), кроме последнего раунда.
            if rnd < job["rounds_total"] and not job["cancel"].is_set():
                target = time.time() + job["interval_sec"]
                with job["lock"]:
                    job["next_run_at"] = target
                while time.time() < target and not job["cancel"].is_set():
                    time.sleep(1)
                with job["lock"]:
                    job["next_run_at"] = None
    except Exception as e:  # noqa: BLE001
        with job["lock"]:
            job["status"] = "error"
            job["fatal"] = str(e)
            job["next_run_at"] = None
        return

    with job["lock"]:
        job["status"] = "cancelled" if job["cancel"].is_set() else "done"
        job["finished_at"] = time.time()
        job["next_run_at"] = None


def monitor_rows(job):
    """Сводка по парам ключ+домен (× источник, если их несколько: Live/XML)."""
    snaps = job["snapshots"]
    sources = job.get("sources") or [{"name": "single", "label": None}]
    multi = len(sources) > 1
    rows = []
    for kw, dom in job.get("pairs", []):
        for s in sources:
            sn = s["name"]
            positions = [snap["positions"].get(sn, {}).get(kw, {}).get(dom) for snap in snaps]
            found = [p for p in positions if p is not None]
            rows.append(
                {
                    "keyword": kw,
                    "domain": dom,
                    "source": s["label"] if multi else None,
                    "positions": positions,
                    "avg": round(sum(found) / len(found), 1) if found else None,
                    "best": min(found) if found else None,
                    "worst": max(found) if found else None,
                    "found": len(found),
                    "checks": len(snaps),
                }
            )
    return rows


# --------------------------------------------------------------------------- #
#  Дата регистрации домена: RDAP → WHOIS, с durable-кэшем                     #
# --------------------------------------------------------------------------- #
# «Настоящий запуск» дорвея нельзя доказать без даты регистрации домена.
# RDAP (JSON) стабильнее классического WHOIS и покрывает все gTLD (.xyz/.top/
# .buzz/.casino/.team/.click/…), а также .ru через бутстрап rdap.org. Один
# запрос на домен за всё окно — результат кэшируется на диск и переживает
# перезапуск сервера (и переиспользуется между прогонами).
WHOIS_CACHE_FILE = os.path.join(OUTPUT_DIR, "whois_cache.json")
_WHOIS_BIN = shutil.which("whois")
_RDAP_URL = "https://rdap.org/domain/"
_WHOIS_DATE_RE = re.compile(
    r"(?:creation date|created on|created|registered on|registration date|"
    r"registration time|domain registration date)\s*:?\s*"
    r"(\d{4}-\d{2}-\d{2}(?:[ tT]\d{2}:\d{2}(?::\d{2})?(?:\.\d+)?(?:z|[+-]\d{2}:?\d{2})?)?"
    r"|\d{2}-[a-zA-Z]{3}-\d{4})",
    re.I,
)


def _load_whois_cache():
    try:
        with open(WHOIS_CACHE_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_whois_cache(cache):
    try:
        tmp = WHOIS_CACHE_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)
        os.replace(tmp, WHOIS_CACHE_FILE)
    except Exception:
        pass


def _rdap_created(regdom, timeout):
    """Дата регистрации через RDAP. Возвращает (iso_str|None, retry_bool)."""
    try:
        r = requests.get(_RDAP_URL + regdom, timeout=timeout,
                         headers={"Accept": "application/rdap+json"},
                         allow_redirects=True)
    except requests.RequestException:
        return None, True
    if r.status_code == 429 or r.status_code >= 500:
        return None, True
    if r.status_code != 200:
        return None, False
    try:
        data = r.json()
    except Exception:
        return None, False
    for ev in (data.get("events") or []):
        if ev.get("eventAction") == "registration":
            d = (ev.get("eventDate") or "").strip()
            if d:
                return d, False
    return None, False


def _whois_created(regdom, timeout):
    """Фолбэк на системный whois (есть в macOS). Возвращает строку-дату или None."""
    if not _WHOIS_BIN:
        return None
    try:
        out = subprocess.run([_WHOIS_BIN, regdom], stdout=subprocess.PIPE,
                             stderr=subprocess.DEVNULL, universal_newlines=True,
                             timeout=timeout).stdout or ""
    except Exception:
        return None
    m = _WHOIS_DATE_RE.search(out)
    return m.group(1).strip() if m else None


def domain_created(regdom, cfg):
    """RDAP с ретраями/бэкоффом, затем whois. Возвращает (created|None, source)."""
    for attempt in range(cfg.get("whois_retries", 1) + 1):
        created, retry = _rdap_created(regdom, cfg.get("whois_timeout", 15))
        if created:
            return created, "rdap"
        if not retry:
            break
        time.sleep(min(2 ** attempt, 6))
    w = _whois_created(regdom, cfg.get("whois_timeout", 15))
    if w:
        return w, "whois"
    return None, ""


# --------------------------------------------------------------------------- #
#  Мониторинг запусков: частые срезы + дата регистрации                       #
# --------------------------------------------------------------------------- #
# Существующие колонки «Сбора ТОП» + 4 новых в конец (совместимо с пайплайном).
WATCH_COLS = ["query", "position", "url", "title", "domain",
              "scrape_ts", "regdom", "whois_created", "whois_source"]


def _watch_build(queries, results_by_query, seen, scrape_ts):
    """Сырьё среза. Возвращает (rows_raw, wide_rows, misses, snap_new, snap_new_set).
    rows_raw: [(query, pos, url, title, host, regdom)]; snap_new — уникальные
    regdom, впервые встреченные в этом окне (для них проставим whois_created).
    misses — запросы, не отработавшие (ошибка/капча/таймаут), НЕ «пусто по делу»."""
    rows_raw, wide_rows, misses = [], [], []
    snap_new, snap_new_set = [], set()
    for q in queries:
        res, err = results_by_query.get(q, ([], None))
        if err and not res:
            misses.append(q)
        for i, r in enumerate(res, 1):
            host = (r.get("domain") or domain_from_url(r.get("url", ""))).lower()
            rd = registrable_domain(host)
            rows_raw.append((q, i, r["url"], r.get("title", ""), host, rd))
            if rd and rd not in seen and rd not in snap_new_set:
                snap_new_set.add(rd)
                snap_new.append(rd)
        wide_rows.append((q, [r["url"] for r in res]))
    return rows_raw, wide_rows, misses, snap_new, snap_new_set


def _write_watch_xlsx(path, rows, wide_rows, top_n):
    try:
        from openpyxl import Workbook
    except Exception:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(WATCH_COLS)
    for row in rows:
        ws.append(row)
    ws2 = wb.create_sheet("wide")
    ws2.append(["query"] + [f"url_{i}" for i in range(1, top_n + 1)])
    for q, urls in wide_rows:
        ws2.append([q] + (list(urls) + [""] * top_n)[:top_n])
    try:
        wb.save(path)
        return True
    except Exception:
        return False


def _write_watch_csv(path, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(WATCH_COLS)
        w.writerows(rows)


def _watch_enrich(job, cfg, snap_new, cache):
    """RDAP/WHOIS для новых regdom этого среза. Возвращает {regdom: (created, source)}
    и обновляет durable-кэш. Считает прогресс в job (whois_done/whois_total)."""
    mapping, todo = {}, []
    for rd in snap_new:
        if rd in cache:
            mapping[rd] = (cache[rd].get("created"), cache[rd].get("source"))
            with job["lock"]:
                job["whois_done"] += 1
        else:
            todo.append(rd)
    clock = Lock()

    def work(rd):
        if job["cancel"].is_set():
            return rd, (None, "")
        created, source = domain_created(rd, cfg)
        if cfg["whois_delay"] > 0:
            time.sleep(cfg["whois_delay"])
        return rd, (created, source)

    if todo:
        with ThreadPoolExecutor(max_workers=cfg["whois_workers"]) as ex:
            futs = {ex.submit(work, rd): rd for rd in todo}
            for fut in as_completed(futs):
                rd, (created, source) = fut.result()
                with clock:
                    mapping[rd] = (created, source)
                    cache[rd] = {"created": created, "source": source}
                with job["lock"]:
                    job["whois_done"] += 1
        _save_whois_cache(cache)
    return mapping


def run_watch(job_id, cfg, queries):
    job = WATCHJOBS[job_id]
    seen = set()                       # regdom, встреченные в этом окне
    cache = _load_whois_cache()
    try:
        for rnd in range(1, job["rounds_total"] + 1):
            if job["cancel"].is_set():
                break
            round_start = time.time()
            scrape_dt = datetime.now().astimezone()
            stamp = scrape_dt.strftime("%Y%m%d_%H%M%S")
            scrape_ts = scrape_dt.isoformat(timespec="seconds")
            with job["lock"]:
                job["phase"] = "сбор выдачи"
                job["cur_done"], job["cur_total"] = 0, len(queries)
                job["whois_done"], job["whois_total"] = 0, 0

            # 1) снять выдачу по всем запросам
            rbq = {}
            with ThreadPoolExecutor(max_workers=cfg["workers"]) as ex:
                futs = {ex.submit(fetch_one, cfg, q, job["cancel"]): q for q in queries}
                for fut in as_completed(futs):
                    q = futs[fut]
                    try:
                        rbq[q] = fut.result()
                    except Exception as e:  # noqa: BLE001
                        rbq[q] = ([], f"исключение: {e}")
                    with job["lock"]:
                        job["cur_done"] += 1
            if job["cancel"].is_set():
                break

            # 2) сырьё + новые regdom этого окна
            rows_raw, wide_rows, misses, snap_new, snap_new_set = _watch_build(
                queries, rbq, seen, scrape_ts)

            # 3) дата регистрации для впервые виденных regdom
            mapping = {}
            if cfg["do_whois"] and snap_new and not job["cancel"].is_set():
                with job["lock"]:
                    job["phase"] = "дата регистрации (RDAP)"
                    job["whois_total"] = len(snap_new)
                mapping = _watch_enrich(job, cfg, snap_new, cache)
            seen.update(snap_new_set)

            # 4) строки: whois_created только у впервые виденных regdom
            long_rows = []
            for (q, i, url, title, host, rd) in rows_raw:
                created, source = mapping.get(rd, (None, "")) if rd in snap_new_set else ("", "")
                long_rows.append([q, i, url, title, host, scrape_ts, rd,
                                  created or "", source or ""])

            # 5) файлы среза: serp_ДАТА_ВРЕМЯ.xlsx/.csv (+ misses_*.txt)
            xlsx_path = os.path.join(OUTPUT_DIR, f"serp_{stamp}.xlsx")
            csv_path = os.path.join(OUTPUT_DIR, f"serp_{stamp}.csv")
            xlsx_ok = _write_watch_xlsx(xlsx_path, long_rows, wide_rows, cfg["top_n"])
            _write_watch_csv(csv_path, long_rows)
            misses_path = None
            if misses:
                misses_path = os.path.join(OUTPUT_DIR, f"misses_{stamp}.txt")
                try:
                    with open(misses_path, "w", encoding="utf-8") as f:
                        f.write("\n".join(misses))
                except Exception:
                    misses_path = None

            fresh = sum(1 for rd in snap_new if mapping.get(rd, (None, ""))[0])
            with job["lock"]:
                job["snapshots"].append({
                    "round": rnd, "at": round_start, "scrape_ts": scrape_ts,
                    "xlsx": os.path.basename(xlsx_path) if xlsx_ok else None,
                    "csv": os.path.basename(csv_path),
                    "xlsx_path": xlsx_path if xlsx_ok else None,
                    "csv_path": csv_path, "misses_path": misses_path,
                    "results": len(long_rows), "misses": len(misses),
                    "new_domains": len(snap_new), "whois_found": fresh,
                })
                job["rounds_done"] = rnd
                job["total_new"] += len(snap_new)
                for rd in snap_new:
                    c, s = mapping.get(rd, (None, ""))
                    if c:
                        job["fresh_list"].append(
                            {"regdom": rd, "created": c, "source": s, "round": rnd})
                job["fresh_list"] = job["fresh_list"][-800:]
                job["phase"] = "пауза"

            # 6) периодическая пауза до следующего среза (прерываемая)
            if rnd < job["rounds_total"] and not job["cancel"].is_set():
                target = round_start + job["interval_sec"]
                with job["lock"]:
                    job["next_run_at"] = target
                while time.time() < target and not job["cancel"].is_set():
                    time.sleep(1)
                with job["lock"]:
                    job["next_run_at"] = None
    except Exception as e:  # noqa: BLE001
        with job["lock"]:
            job["status"] = "error"
            job["fatal"] = str(e)
            job["next_run_at"] = None
        return
    with job["lock"]:
        job["status"] = "cancelled" if job["cancel"].is_set() else "done"
        job["finished_at"] = time.time()
        job["next_run_at"] = None
        job["phase"] = "—"


# --------------------------------------------------------------------------- #
#  HTTP-маршруты                                                              #
# --------------------------------------------------------------------------- #
@app.route("/")
def index():
    resp = Response(render_template("index.html"), mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.route("/api/version")
def api_version():
    return jsonify({"build": APP_BUILD})


@app.route("/api/run", methods=["POST"])
def api_run():
    data = request.get_json(force=True, silent=True) or {}

    user = (data.get("user") or "").strip()
    key = (data.get("key") or "").strip()
    if not user or not key:
        return jsonify({"error": "Укажите User ID и API key из кабинета xmlstock."}), 400

    raw = data.get("queries") or ""
    seen, queries = set(), []
    for line in raw.replace("\r", "\n").split("\n"):
        q = line.strip()
        if q and q not in seen:
            seen.add(q)
            queries.append(q)
    if not queries:
        return jsonify({"error": "Список запросов пуст."}), 400

    engine = data.get("engine") or "yandex"
    endpoint = (data.get("endpoint") or "").strip() or ENDPOINTS.get(engine)
    if not endpoint or not endpoint.lower().startswith(("http://", "https://")):
        return jsonify({"error": "Некорректный URL эндпоинта."}), 400

    def clamp(val, lo, hi, default):
        try:
            return max(lo, min(hi, int(val)))
        except (TypeError, ValueError):
            return default

    def fnum(val, default):
        try:
            return max(0.0, float(val))
        except (TypeError, ValueError):
            return default

    cfg = {
        "user": user,
        "key": key,
        "endpoint": endpoint,
        "live": _is_live(endpoint),
        "lr": (data.get("lr") or "").strip(),
        "domain": (data.get("domain") or "").strip(),
        "device": (data.get("device") or "").strip(),
        "top_n": clamp(data.get("top_n"), 1, 50, 10),
        "workers": clamp(data.get("workers"), 1, 20, 5),
        "delay": fnum(data.get("delay"), 0.0),
        "timeout": clamp(data.get("timeout"), 5, 120, 30),
        "retries": clamp(data.get("retries"), 0, 5, 2),
    }

    job_id = uuid.uuid4().hex
    with JOBS_LOCK:
        JOBS[job_id] = {
            "id": job_id,
            "status": "running",
            "total": len(queries),
            "done": 0,
            "errors": [],
            "recent": [],
            "lock": Lock(),
            "cancel": Event(),
            "csv_path": None,
            "xlsx_path": None,
            "rows": 0,
            "result_urls": [],
            "started_at": time.time(),
        }
        LAST_COLLECT["id"] = job_id

    Thread(target=run_job, args=(job_id, cfg, queries), daemon=True).start()
    return jsonify({"job_id": job_id, "total": len(queries), "deduped": len(queries)})


@app.route("/api/status/<job_id>")
def api_status(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    with job["lock"]:
        return jsonify(
            {
                "status": job["status"],
                "total": job["total"],
                "done": job["done"],
                "errors_count": len(job["errors"]),
                "recent": job["recent"],
                "rows": job.get("rows", 0),
                "has_csv": bool(job.get("csv_path")),
                "has_xlsx": bool(job.get("xlsx_path")),
                "fatal": job.get("fatal"),
            }
        )


@app.route("/api/cancel/<job_id>", methods=["POST"])
def api_cancel(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    job["cancel"].set()
    return jsonify({"ok": True})


@app.route("/api/download/<job_id>")
def api_download(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    fmt = request.args.get("fmt", "xlsx")
    path = job.get("xlsx_path") if fmt == "xlsx" else job.get("csv_path")
    if not path or not os.path.exists(path):
        return jsonify({"error": "Файл ещё не готов"}), 404
    return send_file(path, as_attachment=True, download_name=os.path.basename(path))


@app.route("/api/errors/<job_id>")
def api_errors(job_id):
    job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    with job["lock"]:
        return jsonify({"errors": job["errors"]})


# --- Трекер позиций -------------------------------------------------------- #
def _split_unique(raw, limit):
    seen, out = set(), []
    for line in (raw or "").replace("\r", "\n").split("\n"):
        v = line.strip()
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out[:limit]


def _split_multi(raw):
    """Домены в одном поле через запятую/пробел/перенос строки."""
    s = (raw or "").replace(",", " ").replace(";", " ").replace("\r", " ").replace("\n", " ")
    return [p for p in s.split() if p]


@app.route("/api/monitor/run", methods=["POST"])
def api_monitor_run():
    data = request.get_json(force=True, silent=True) or {}

    user = (data.get("user") or "").strip()
    key = (data.get("key") or "").strip()
    if not user or not key:
        return jsonify({"error": "Укажите User ID и API key из кабинета xmlstock."}), 400

    # Блоки «домен(ы) + его ключи». Старый формат (общие keywords+domains)
    # поддерживается как один блок ради совместимости.
    def parse_group(g):
        kws = _split_unique((g or {}).get("keywords"), 100)
        doms, dseen = [], set()
        for d in _split_multi((g or {}).get("domains")):
            nd = normalize_domain(d)
            if nd and nd not in dseen:
                dseen.add(nd)
                doms.append(nd)
        return {"domains": doms, "keywords": kws} if (kws and doms) else None

    raw_groups = data.get("groups")
    groups = []
    if isinstance(raw_groups, list) and raw_groups:
        for g in raw_groups[:20]:
            parsed = parse_group(g)
            if parsed:
                groups.append(parsed)
    else:
        parsed = parse_group({"keywords": data.get("keywords"), "domains": data.get("domains")})
        if parsed:
            groups.append(parsed)

    if not groups:
        return jsonify({"error": "Добавьте хотя бы один домен с его ключами."}), 400

    # Уникальные ключи (запрашиваются 1 раз за раунд) и пары (ключ, домен).
    keywords, kseen = [], set()
    pairs, pairs_by_kw, pseen = [], {}, set()
    for g in groups:
        for kw in g["keywords"]:
            if kw not in kseen:
                kseen.add(kw)
                keywords.append(kw)
        for dom in g["domains"]:
            for kw in g["keywords"]:
                if (kw, dom) not in pseen:
                    pseen.add((kw, dom))
                    pairs.append([kw, dom])
                pairs_by_kw.setdefault(kw, [])
                if dom not in pairs_by_kw[kw]:
                    pairs_by_kw[kw].append(dom)
    keywords = keywords[:100]
    all_domains = list(dict.fromkeys(d for g in groups for d in g["domains"]))

    engine = data.get("engine") or "yandex"
    custom = (data.get("endpoint") or "").strip()
    both = bool(data.get("both"))
    family = "google" if str(engine).startswith("google") else "yandex"

    # Источники съёма. «both» — снять Live и XML одновременно (для сравнения).
    if both:
        pair = ("google_live", "google") if family == "google" else ("yandex_live", "yandex")
        sources = [
            {"name": "live", "label": "Live", "endpoint": ENDPOINTS[pair[0]], "live": True},
            {"name": "xml", "label": "XML", "endpoint": ENDPOINTS[pair[1]], "live": False},
        ]
    else:
        endpoint = custom or ENDPOINTS.get(engine)
        if not endpoint or not endpoint.lower().startswith(("http://", "https://")):
            return jsonify({"error": "Некорректный URL эндпоинта."}), 400
        live = _is_live(endpoint)
        sources = [{"name": "single", "label": "Live" if live else "XML",
                    "endpoint": endpoint, "live": live}]

    def clamp(val, lo, hi, default):
        try:
            return max(lo, min(hi, int(val)))
        except (TypeError, ValueError):
            return default

    try:
        interval_min = float(data.get("interval_min"))
    except (TypeError, ValueError):
        interval_min = 4.0
    interval_sec = int(max(5, min(86400, interval_min * 60)))  # до 24 часов
    rounds_total = clamp(data.get("rounds"), 1, 100, 10)

    cfg_base = {
        "user": user,
        "key": key,
        "lr": (data.get("lr") or "").strip(),
        "domain": (data.get("domain") or "").strip(),
        "device": (data.get("device") or "").strip(),
        "top_n": clamp(data.get("depth"), 10, 100, 100),
        "workers": clamp(data.get("workers"), 1, 10, 5),
        "delay": 0.0,
        "timeout": 30,
        "retries": 2,
    }

    job_id = uuid.uuid4().hex
    with MON_LOCK:
        MONITORS[job_id] = {
            "id": job_id,
            "status": "running",
            "keywords": keywords,
            "domains": all_domains,
            "groups": groups,
            "pairs": pairs,
            "pairs_by_kw": pairs_by_kw,
            "sources": sources,
            "rounds_total": rounds_total,
            "rounds_done": 0,
            "interval_sec": interval_sec,
            "depth": cfg_base["top_n"],
            "snapshots": [],
            "errors": [],
            "next_run_at": None,
            "lock": Lock(),
            "cancel": Event(),
            "started_at": time.time(),
        }
        ACTIVE_MONITOR["id"] = job_id

    Thread(target=run_monitor, args=(job_id, cfg_base), daemon=True).start()
    return jsonify(
        {
            "job_id": job_id,
            "keywords": len(keywords),
            "domains": all_domains,
            "pairs": len(pairs),
            "sources": [s["label"] for s in sources],
            "multi": len(sources) > 1,
            "rounds": rounds_total,
            "interval_sec": interval_sec,
            "depth": cfg_base["top_n"],
        }
    )


@app.route("/api/monitor/status/<job_id>")
def api_monitor_status(job_id):
    job = MONITORS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    with job["lock"]:
        nxt = job.get("next_run_at")
        return jsonify(
            {
                "status": job["status"],
                "rounds_total": job["rounds_total"],
                "rounds_done": job["rounds_done"],
                "interval_sec": job["interval_sec"],
                "seconds_to_next": int(max(0, nxt - time.time())) if nxt else None,
                "depth": job["depth"],
                "domains": job["domains"],
                "multi": len(job.get("sources", [])) > 1,
                "rows": monitor_rows(job),
                "errors_count": len(job["errors"]),
                "fatal": job.get("fatal"),
            }
        )


@app.route("/api/monitor/cancel/<job_id>", methods=["POST"])
def api_monitor_cancel(job_id):
    job = MONITORS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    job["cancel"].set()
    return jsonify({"ok": True})


@app.route("/api/monitor/active")
def api_monitor_active():
    jid = ACTIVE_MONITOR.get("id")
    if jid and jid in MONITORS:
        return jsonify({"job_id": jid, "status": MONITORS[jid]["status"]})
    return jsonify({"job_id": None})


@app.route("/api/monitor/download/<job_id>")
def api_monitor_download(job_id):
    job = MONITORS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    fmt = request.args.get("fmt", "csv")
    with job["lock"]:
        rows = monitor_rows(job)
        n = len(job["snapshots"])
        multi = len(job.get("sources", [])) > 1
    src_head = ["source"] if multi else []
    headers = (
        ["keyword", "domain"] + src_head
        + ["avg_position", "best", "worst", "found", "checks"]
        + [f"round_{i}" for i in range(1, n + 1)]
    )

    def cells(r):
        return (
            [r["keyword"], r["domain"]]
            + ([r["source"]] if multi else [])
            + [r["avg"], r["best"], r["worst"], r["found"], r["checks"]]
            + [p if p is not None else "" for p in r["positions"]]
        )

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            return jsonify({"error": "openpyxl не установлен"}), 400
        wb = Workbook()
        ws = wb.active
        ws.title = "positions"
        ws.append(headers)
        for r in rows:
            ws.append(cells(r))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"positions_{job_id[:8]}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    sbuf = io.StringIO()
    w = csv.writer(sbuf, delimiter=";")
    w.writerow(headers)
    for r in rows:
        w.writerow(cells(r))
    payload = ("﻿" + sbuf.getvalue()).encode("utf-8")
    return send_file(
        io.BytesIO(payload),
        as_attachment=True,
        download_name=f"positions_{job_id[:8]}.csv",
        mimetype="text/csv",
    )


# --- Сбор страниц сайта (site:) -------------------------------------------- #
@app.route("/api/site/run", methods=["POST"])
def api_site_run():
    data = request.get_json(force=True, silent=True) or {}
    user = (data.get("user") or "").strip()
    key = (data.get("key") or "").strip()
    if not user or not key:
        return jsonify({"error": "Укажите User ID и API key из кабинета xmlstock."}), 400

    domains = [d for d in _split_unique(data.get("domains"), 200) if clean_host(d)]
    if not domains:
        return jsonify({"error": "Добавьте хотя бы один домен."}), 400

    engine = data.get("engine") or "yandex"
    custom = (data.get("endpoint") or "").strip()
    endpoint = custom or ENDPOINTS.get(engine)
    if not endpoint or not endpoint.lower().startswith(("http://", "https://")):
        return jsonify({"error": "Некорректный URL эндпоинта."}), 400
    # site: полнее и дешевле на XML — если выбран Live (без своего URL), берём XML.
    if not custom and _is_live(endpoint):
        endpoint = ENDPOINTS["google"] if str(engine).startswith("google") else ENDPOINTS["yandex"]

    op = data.get("operator") or "site:"
    if op not in ("site:", "host:"):
        op = "site:"

    def clamp(val, lo, hi, default):
        try:
            return max(lo, min(hi, int(val)))
        except (TypeError, ValueError):
            return default

    try:
        delay = max(0.0, min(5.0, float(data.get("delay"))))
    except (TypeError, ValueError):
        delay = 0.0
    cfg = {
        "user": user, "key": key, "endpoint": endpoint, "live": _is_live(endpoint),
        "operator": op,
        "deep": bool(data.get("deep", False)),
        "multi_op": bool(data.get("multi_op", False)),
        "deep_subdomains": clamp(data.get("subdomains"), 0, 500, 150),
        "lr": (data.get("lr") or "").strip(),
        "device": (data.get("device") or "").strip(),
        "max_urls": clamp(data.get("max_urls"), 10, 5000, 1000),
        "workers": clamp(data.get("workers"), 1, 10, 4),
        "delay": delay, "timeout": 30, "retries": 2,
    }

    job_id = uuid.uuid4().hex
    with SITE_LOCK:
        SITEJOBS[job_id] = {
            "id": job_id, "status": "running", "domains": domains,
            "total": len(domains), "done": 0, "results": {}, "errors": [],
            "lock": Lock(), "cancel": Event(), "started_at": time.time(),
        }
        ACTIVE_SITE["id"] = job_id

    Thread(target=run_site, args=(job_id, cfg, domains), daemon=True).start()
    return jsonify({"job_id": job_id, "total": len(domains),
                    "operator": op, "max_urls": cfg["max_urls"]})


@app.route("/api/site/status/<job_id>")
def api_site_status(job_id):
    job = SITEJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    with job["lock"]:
        rows, total_urls = [], 0
        for d in job["domains"]:
            r = job["results"].get(d)
            if r:
                total_urls += r["collected"]
                rows.append({"domain": d, "found": r.get("found"),
                             "collected": r["collected"], "error": r.get("error"),
                             "per_op": r.get("per_op")})
            else:
                rows.append({"domain": d, "found": None, "collected": None, "error": None})
        return jsonify({"status": job["status"], "total": job["total"], "done": job["done"],
                        "rows": rows, "total_urls": total_urls,
                        "errors_count": len(job["errors"]), "fatal": job.get("fatal")})


@app.route("/api/site/cancel/<job_id>", methods=["POST"])
def api_site_cancel(job_id):
    job = SITEJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    job["cancel"].set()
    return jsonify({"ok": True})


@app.route("/api/site/active")
def api_site_active():
    jid = ACTIVE_SITE.get("id")
    if jid and jid in SITEJOBS:
        return jsonify({"job_id": jid, "status": SITEJOBS[jid]["status"]})
    return jsonify({"job_id": None})


@app.route("/api/site/urls/<job_id>")
def api_site_urls(job_id):
    job = SITEJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    limit, out = 5000, []
    with job["lock"]:
        for d in job["domains"]:
            r = job["results"].get(d)
            if not r:
                continue
            for u in r["urls"]:
                out.append({"domain": d, "url": u["url"], "title": u["title"]})
                if len(out) >= limit:
                    break
            if len(out) >= limit:
                break
    return jsonify({"urls": out, "truncated": len(out) >= limit})


@app.route("/api/site/download/<job_id>")
def api_site_download(job_id):
    job = SITEJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    fmt = request.args.get("fmt", "csv")
    with job["lock"]:
        domains = list(job["domains"])
        results = {d: job["results"].get(d) for d in domains}

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            return jsonify({"error": "openpyxl не установлен"}), 400
        wb = Workbook()
        ws = wb.active
        ws.title = "urls"
        ws.append(["domain", "url", "title"])
        for d in domains:
            r = results.get(d)
            if not r:
                continue
            for u in r["urls"]:
                ws.append([d, u["url"], u["title"]])
        ws2 = wb.create_sheet("summary")
        ws2.append(["domain", "found_estimate", "collected"])
        for d in domains:
            r = results.get(d) or {}
            ws2.append([d, r.get("found"), r.get("collected", 0)])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"site_pages_{job_id[:8]}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    sbuf = io.StringIO()
    w = csv.writer(sbuf, delimiter=";")
    w.writerow(["domain", "url", "title"])
    for d in domains:
        r = results.get(d)
        if not r:
            continue
        for u in r["urls"]:
            w.writerow([d, u["url"], u["title"]])
    payload = ("﻿" + sbuf.getvalue()).encode("utf-8")
    return send_file(io.BytesIO(payload), as_attachment=True,
                     download_name=f"site_pages_{job_id[:8]}.csv", mimetype="text/csv")


# --- Анализ редиректов ----------------------------------------------------- #
_UA = {
    "mobile": ("Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 "
               "(KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1"),
    "desktop": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"),
}


@app.route("/api/collect/last_urls")
def api_collect_last_urls():
    snap = _load_last_collect()
    if not snap:
        return jsonify({"urls": [], "status": None})
    return jsonify({"urls": snap.get("urls", []), "status": "done"})


@app.route("/api/collect/last")
def api_collect_last():
    snap = _load_last_collect()
    if not snap:
        return jsonify({"exists": False})
    return jsonify({
        "exists": True, "at": snap.get("at"), "total": snap.get("total"),
        "rows": snap.get("rows"), "urls_count": len(snap.get("urls", [])),
        "has_csv": bool(snap.get("csv") and os.path.exists(snap["csv"])),
        "has_xlsx": bool(snap.get("xlsx") and os.path.exists(snap["xlsx"])),
    })


@app.route("/api/collect/last_download")
def api_collect_last_download():
    snap = _load_last_collect()
    if not snap:
        return jsonify({"error": "Нет последнего сбора"}), 404
    fmt = request.args.get("fmt", "csv")
    path = snap.get("xlsx") if fmt == "xlsx" else snap.get("csv")
    if not path or not os.path.exists(path):
        return jsonify({"error": "Файл не найден"}), 404
    return send_file(path, as_attachment=True, download_name=os.path.basename(path))


@app.route("/api/redirects/run", methods=["POST"])
def api_redirects_run():
    data = request.get_json(force=True, silent=True) or {}
    seen, urls = set(), []
    for line in (data.get("urls") or "").replace("\r", "\n").split("\n"):
        u = line.strip()
        if u and u not in seen:
            seen.add(u)
            urls.append(u)
    if not urls:
        return jsonify({"error": "Добавьте ссылки (или загрузите из «Сбора ТОП»)."}), 400

    def clamp(val, lo, hi, default):
        try:
            return max(lo, min(hi, int(val)))
        except (TypeError, ValueError):
            return default

    ua_key = data.get("ua") if data.get("ua") in _UA else "mobile"
    cfg = {
        "group_by": "root" if (data.get("group_by") or "root") == "root" else "host",
        "only_subdomains": bool(data.get("only_subdomains", True)),
        "engine_mode": "browser" if data.get("engine_mode") == "browser" else "http",
        "wait_ms": clamp(data.get("wait_ms"), 500, 20000, 5000),
        "click": bool(data.get("click", False)),
        "ua": _UA[ua_key],
        "referer": "https://yandex.ru/" if data.get("referer", True) else "",
        "verify": bool(data.get("verify", False)),
        "timeout": clamp(data.get("timeout"), 3, 60, 15),
        "workers": clamp(data.get("workers"), 1, 30, 10),
    }

    job_id = uuid.uuid4().hex
    with REDIR_LOCK:
        REDIRJOBS[job_id] = {
            "id": job_id, "status": "running", "total": 0, "done": 0,
            "rows": [], "lock": Lock(), "cancel": Event(), "started_at": time.time(),
        }
        ACTIVE_REDIR["id"] = job_id

    Thread(target=run_redirects, args=(job_id, cfg, urls), daemon=True).start()
    return jsonify({"job_id": job_id, "input": len(urls), "group_by": cfg["group_by"]})


@app.route("/api/redirects/status/<job_id>")
def api_redirects_status(job_id):
    job = REDIRJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    with job["lock"]:
        rows = list(job["rows"])
        return jsonify({"status": job["status"], "total": job["total"], "done": job["done"],
                        "rows": rows, "summary": redir_summary(rows), "fatal": job.get("fatal")})


@app.route("/api/redirects/cancel/<job_id>", methods=["POST"])
def api_redirects_cancel(job_id):
    job = REDIRJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    job["cancel"].set()
    return jsonify({"ok": True})


@app.route("/api/redirects/active")
def api_redirects_active():
    jid = ACTIVE_REDIR.get("id")
    if jid and jid in REDIRJOBS:
        return jsonify({"job_id": jid, "status": REDIRJOBS[jid]["status"]})
    return jsonify({"job_id": None})


@app.route("/api/redirects/download/<job_id>")
def api_redirects_download(job_id):
    job = REDIRJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    fmt = request.args.get("fmt", "csv")
    with job["lock"]:
        rows = list(job["rows"])
        summary = redir_summary(rows)
    dcols = ["source_host", "source_url", "redirected", "final_domain", "final_url", "status", "error"]

    def drow(r):
        return [r["source_host"], r["source_url"], "да" if r["redirected"] else "нет",
                r["final_domain"], r["final_url"], r["status"], r["error"] or ""]

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            return jsonify({"error": "openpyxl не установлен"}), 400
        wb = Workbook()
        ws = wb.active
        ws.title = "redirects"
        ws.append(dcols)
        for r in rows:
            ws.append(drow(r))
        ws2 = wb.create_sheet("summary")
        ws2.append(["target_domain", "sites", "percent"])
        for t in summary["targets"]:
            ws2.append([t["domain"], t["count"], t["percent"]])
        ws2.append(["(без редиректа)", summary["no_redirect"], summary["no_redirect_percent"]])
        ws2.append(["(ошибки)", summary["errors"], summary["errors_percent"]])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"redirects_{job_id[:8]}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    sbuf = io.StringIO()
    w = csv.writer(sbuf, delimiter=";")
    w.writerow(dcols)
    for r in rows:
        w.writerow(drow(r))
    w.writerow([])
    w.writerow(["target_domain", "sites", "percent"])
    for t in summary["targets"]:
        w.writerow([t["domain"], t["count"], t["percent"]])
    w.writerow(["(без редиректа)", summary["no_redirect"], summary["no_redirect_percent"]])
    w.writerow(["(ошибки)", summary["errors"], summary["errors_percent"]])
    payload = ("﻿" + sbuf.getvalue()).encode("utf-8")
    return send_file(io.BytesIO(payload), as_attachment=True,
                     download_name=f"redirects_{job_id[:8]}.csv", mimetype="text/csv")


@app.route("/api/collect/last_domains")
def api_collect_last_domains():
    """Уникальные хосты (домены с поддоменами) из последнего «Сбора ТОП»."""
    snap = _load_last_collect()
    if not snap:
        return jsonify({"domains": []})
    seen, out = set(), []
    for u in snap.get("urls", []):
        h = url_host(u)
        if h and h not in seen:
            seen.add(h)
            out.append(h)
    return jsonify({"domains": out})


# --- Проверка живости доменов --------------------------------------------- #
@app.route("/api/live/run", methods=["POST"])
def api_live_run():
    data = request.get_json(force=True, silent=True) or {}
    seen, domains = set(), []
    for line in (data.get("domains") or "").replace("\r", "\n").split("\n"):
        h = url_host(line.strip()) or clean_host(line.strip())
        if h and h not in seen:
            seen.add(h)
            domains.append(h)
    if not domains:
        return jsonify({"error": "Добавьте домены (или загрузите из «Сбора ТОП»)."}), 400

    def clamp(val, lo, hi, default):
        try:
            return max(lo, min(hi, int(val)))
        except (TypeError, ValueError):
            return default

    cfg = {
        "ua": _UA.get(data.get("ua"), _UA["desktop"]),
        "timeout": clamp(data.get("timeout"), 2, 30, 8),
        "workers": clamp(data.get("workers"), 1, 40, 15),
    }
    job_id = uuid.uuid4().hex
    with LIVE_LOCK:
        LIVEJOBS[job_id] = {
            "id": job_id, "status": "running", "total": len(domains), "done": 0,
            "rows": [], "lock": Lock(), "cancel": Event(), "started_at": time.time(),
        }
        ACTIVE_LIVE["id"] = job_id
    Thread(target=run_live, args=(job_id, cfg, domains), daemon=True).start()
    return jsonify({"job_id": job_id, "total": len(domains)})


@app.route("/api/live/status/<job_id>")
def api_live_status(job_id):
    job = LIVEJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    with job["lock"]:
        rows = list(job["rows"])
        alive = sum(1 for r in rows if r["alive"])
        return jsonify({"status": job["status"], "total": job["total"], "done": job["done"],
                        "rows": rows, "alive": alive, "dead": len(rows) - alive,
                        "fatal": job.get("fatal")})


@app.route("/api/live/cancel/<job_id>", methods=["POST"])
def api_live_cancel(job_id):
    job = LIVEJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    job["cancel"].set()
    return jsonify({"ok": True})


@app.route("/api/live/active")
def api_live_active():
    jid = ACTIVE_LIVE.get("id")
    if jid and jid in LIVEJOBS:
        return jsonify({"job_id": jid, "status": LIVEJOBS[jid]["status"]})
    return jsonify({"job_id": None})


@app.route("/api/live/download/<job_id>")
def api_live_download(job_id):
    job = LIVEJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    fmt = request.args.get("fmt", "csv")
    with job["lock"]:
        rows = list(job["rows"])

    if fmt == "alive":  # только живые — списком, по одному в строке
        payload = ("﻿" + "\n".join(r["host"] for r in rows if r["alive"])).encode("utf-8")
        return send_file(io.BytesIO(payload), as_attachment=True,
                         download_name=f"alive_{job_id[:8]}.txt", mimetype="text/plain")

    cols = ["domain", "alive", "status", "note", "final_url", "error"]

    def cells(r):
        return [r["host"], "жив" if r["alive"] else "мёртв", r["status"] or "",
                r["note"], r["final_url"], r["error"] or ""]

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            return jsonify({"error": "openpyxl не установлен"}), 400
        wb = Workbook()
        ws = wb.active
        ws.title = "liveness"
        ws.append(cols)
        for r in rows:
            ws.append(cells(r))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"liveness_{job_id[:8]}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    sbuf = io.StringIO()
    w = csv.writer(sbuf, delimiter=";")
    w.writerow(cols)
    for r in rows:
        w.writerow(cells(r))
    payload = ("﻿" + sbuf.getvalue()).encode("utf-8")
    return send_file(io.BytesIO(payload), as_attachment=True,
                     download_name=f"liveness_{job_id[:8]}.csv", mimetype="text/csv")


# --- Мониторинг запусков (частые срезы + дата регистрации) ----------------- #
@app.route("/api/watch/run", methods=["POST"])
def api_watch_run():
    data = request.get_json(force=True, silent=True) or {}
    user = (data.get("user") or "").strip()
    key = (data.get("key") or "").strip()
    if not user or not key:
        return jsonify({"error": "Укажите User ID и API key из кабинета xmlstock."}), 400

    raw = data.get("queries") or ""
    seen, queries = set(), []
    for line in raw.replace("\r", "\n").split("\n"):
        q = line.strip()
        if q and q not in seen:
            seen.add(q)
            queries.append(q)
    if not queries:
        return jsonify({"error": "Список запросов пуст."}), 400

    engine = data.get("engine") or "yandex"
    endpoint = (data.get("endpoint") or "").strip() or ENDPOINTS.get(engine)
    if not endpoint or not endpoint.lower().startswith(("http://", "https://")):
        return jsonify({"error": "Некорректный URL эндпоинта."}), 400

    def clamp(val, lo, hi, default):
        try:
            return max(lo, min(hi, int(val)))
        except (TypeError, ValueError):
            return default

    def fnum(val, default):
        try:
            return max(0.0, float(val))
        except (TypeError, ValueError):
            return default

    try:
        interval_val = float(data.get("interval"))
    except (TypeError, ValueError):
        interval_val = 4.0
    unit = data.get("interval_unit") or "hour"
    interval_sec = int(max(10, min(86400, interval_val * (3600 if unit == "hour" else 60))))
    rounds_total = clamp(data.get("rounds"), 1, 200, 12)

    cfg = {
        "user": user, "key": key, "endpoint": endpoint, "live": _is_live(endpoint),
        "lr": (data.get("lr") or "").strip(),
        "domain": (data.get("domain") or "").strip(),
        "device": (data.get("device") or "").strip(),
        "top_n": clamp(data.get("depth"), 1, 50, 10),
        "workers": clamp(data.get("workers"), 1, 20, 5),
        "delay": fnum(data.get("delay"), 0.0),
        "timeout": 30, "retries": 2,
        "do_whois": bool(data.get("do_whois", True)),
        "whois_workers": clamp(data.get("whois_workers"), 1, 10, 4),
        "whois_delay": fnum(data.get("whois_delay"), 0.3),
        "whois_timeout": clamp(data.get("whois_timeout"), 5, 60, 15),
        "whois_retries": clamp(data.get("whois_retries"), 0, 3, 1),
    }

    job_id = uuid.uuid4().hex
    with WATCH_LOCK:
        WATCHJOBS[job_id] = {
            "id": job_id, "status": "running", "total_queries": len(queries),
            "rounds_total": rounds_total, "rounds_done": 0,
            "interval_sec": interval_sec, "depth": cfg["top_n"],
            "do_whois": cfg["do_whois"], "phase": "старт",
            "cur_done": 0, "cur_total": len(queries),
            "whois_done": 0, "whois_total": 0,
            "snapshots": [], "fresh_list": [], "total_new": 0,
            "next_run_at": None, "lock": Lock(), "cancel": Event(),
            "started_at": time.time(),
        }
        ACTIVE_WATCH["id"] = job_id
    Thread(target=run_watch, args=(job_id, cfg, queries), daemon=True).start()
    return jsonify({"job_id": job_id, "queries": len(queries),
                    "rounds": rounds_total, "interval_sec": interval_sec,
                    "whois": cfg["do_whois"], "whois_cli": bool(_WHOIS_BIN)})


@app.route("/api/watch/status/<job_id>")
def api_watch_status(job_id):
    job = WATCHJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    with job["lock"]:
        nxt = job.get("next_run_at")
        snaps = [{"round": s["round"], "scrape_ts": s["scrape_ts"],
                  "xlsx": s["xlsx"], "csv": s["csv"],
                  "results": s["results"], "misses": s["misses"],
                  "new_domains": s["new_domains"], "whois_found": s["whois_found"],
                  "has_misses": bool(s.get("misses_path"))}
                 for s in job["snapshots"]]
        return jsonify({
            "status": job["status"], "phase": job.get("phase"),
            "rounds_total": job["rounds_total"], "rounds_done": job["rounds_done"],
            "interval_sec": job["interval_sec"],
            "seconds_to_next": int(max(0, nxt - time.time())) if nxt else None,
            "cur_done": job["cur_done"], "cur_total": job["cur_total"],
            "whois_done": job["whois_done"], "whois_total": job["whois_total"],
            "depth": job["depth"], "total_new": job.get("total_new", 0),
            "snapshots": snaps, "fresh": job["fresh_list"][-60:][::-1],
            "fatal": job.get("fatal"),
        })


@app.route("/api/watch/cancel/<job_id>", methods=["POST"])
def api_watch_cancel(job_id):
    job = WATCHJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    job["cancel"].set()
    return jsonify({"ok": True})


@app.route("/api/watch/active")
def api_watch_active():
    jid = ACTIVE_WATCH.get("id")
    if jid and jid in WATCHJOBS:
        return jsonify({"job_id": jid, "status": WATCHJOBS[jid]["status"]})
    return jsonify({"job_id": None})


@app.route("/api/watch/file/<job_id>")
def api_watch_file(job_id):
    job = WATCHJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    name = request.args.get("name", "")
    with job["lock"]:
        allowed = {}
        for s in job["snapshots"]:
            if s.get("xlsx"):
                allowed[s["xlsx"]] = s["xlsx_path"]
            if s.get("csv"):
                allowed[s["csv"]] = s["csv_path"]
            if s.get("misses_path"):
                allowed[os.path.basename(s["misses_path"])] = s["misses_path"]
    path = allowed.get(name)
    if not path or not os.path.exists(path):
        return jsonify({"error": "Файл не найден"}), 404
    return send_file(path, as_attachment=True, download_name=name)


@app.route("/api/watch/download_zip/<job_id>")
def api_watch_download_zip(job_id):
    job = WATCHJOBS.get(job_id)
    if not job:
        return jsonify({"error": "Задача не найдена"}), 404
    fmt = "csv" if request.args.get("fmt") == "csv" else "xlsx"
    with job["lock"]:
        snaps = list(job["snapshots"])
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for s in snaps:
            p = s.get("xlsx_path") if fmt == "xlsx" else s.get("csv_path")
            if p and os.path.exists(p):
                z.write(p, os.path.basename(p))
            mp = s.get("misses_path")
            if mp and os.path.exists(mp):
                z.write(mp, os.path.basename(mp))
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"watch_{job_id[:8]}_{fmt}.zip",
                     mimetype="application/zip")


# --------------------------------------------------------------------------- #
#  Журнал запусков: пул доменов + ручные съёмы позиций + отчёт по листам       #
# --------------------------------------------------------------------------- #
# Персистентный «журнал»: несколько запусков (пулов доменов) живут параллельно,
# каждый снимается вручную сколько нужно, всё копится и сохраняется на диск.
# Ключи общие (фиксированные) на все запуски. Съём — Live или XML (переключатель).
LAUNCHES_FILE = os.path.join(OUTPUT_DIR, "launches.json")
LAUNCH_LOCK = Lock()
LAUNCHES = {"keywords": [], "launches": {}}   # загружается при старте
SNAP_PROGRESS = {}                            # launch_id -> прогресс текущего съёма
_SHEET_BAD = re.compile(r"[\[\]:*?/\\]")


def _load_launches():
    global LAUNCHES
    try:
        with open(LAUNCHES_FILE, encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict) and isinstance(data.get("launches"), dict):
            data.setdefault("keywords", [])
            LAUNCHES = data
    except Exception:
        pass


def _save_launches():
    try:
        tmp = LAUNCHES_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(LAUNCHES, f, ensure_ascii=False)
        os.replace(tmp, LAUNCHES_FILE)
    except Exception:
        pass


_load_launches()


def _fmt_ts(ts):
    try:
        return datetime.fromtimestamp(ts).strftime("%d.%m %H:%M")
    except Exception:
        return ""


def _sheet_name(name, used):
    t = _SHEET_BAD.sub(" ", (name or "").strip()) or "Запуск"
    t = t[:31]
    base, i = t, 2
    while t.lower() in used:
        suf = f" ({i})"
        t = base[: 31 - len(suf)] + suf
        i += 1
    used.add(t.lower())
    return t


def _launch_keywords(la):
    """Порядок ключей запуска: объединение по всем съёмам (ключи могли меняться)."""
    out = []
    for s in la.get("snapshots", []):
        for kw in (s.get("keywords") or list(s.get("positions", {}).keys())):
            if kw not in out:
                out.append(kw)
    return out


def launch_report_rows(la, global_keywords=None):
    """Строки отчёта. Ключи = текущие общие (в их порядке) + любые исторические из
    съёмов, которых уже нет в общем списке. Так свежедобавленные ключи видны сразу
    (с прочерками, пока не пересняли), а удалённые из общих — не теряют историю."""
    snaps = la.get("snapshots", [])
    kws = list(global_keywords or [])
    for kw in _launch_keywords(la):
        if kw not in kws:
            kws.append(kw)
    rows = []
    for dom in la["domains"]:
        for kw in kws:
            seq = [s.get("positions", {}).get(kw, {}).get(dom) for s in snaps]
            useq = [s.get("urls", {}).get(kw, {}).get(dom) for s in snaps]
            found = [p for p in seq if p is not None]
            rows.append({
                "domain": dom, "keyword": kw, "positions": seq, "urls": useq,
                "avg": round(sum(found) / len(found), 1) if found else None,
                "best": min(found) if found else None,
                "worst": max(found) if found else None,
                "found": len(found), "checks": len(snaps),
            })
    return rows


def _launch_snapshot(launch_id, cfg, keywords, domains):
    prog = SNAP_PROGRESS[launch_id]
    positions, urls = {}, {}
    try:
        with ThreadPoolExecutor(max_workers=cfg["workers"]) as ex:
            futs = {ex.submit(fetch_one, cfg, kw, prog["cancel"]): kw for kw in keywords}
            for fut in as_completed(futs):
                kw = futs[fut]
                try:
                    results, _ = fut.result()
                except Exception:  # noqa: BLE001
                    results = []
                pmap, umap = {}, {}
                for dom in domains:
                    pos, url = find_position_url(results, dom)
                    pmap[dom] = pos
                    if url:
                        umap[dom] = url          # URL храним только у найденных
                positions[kw] = pmap
                urls[kw] = umap
                with LAUNCH_LOCK:
                    prog["done"] += 1
    except Exception:  # noqa: BLE001
        pass
    snap = {"at": time.time(), "engine": cfg.get("engine_label"),
            "depth": cfg["top_n"], "keywords": list(keywords),
            "positions": positions, "urls": urls}
    with LAUNCH_LOCK:
        la = LAUNCHES["launches"].get(launch_id)
        if la is not None:
            la["snapshots"].append(snap)
            _save_launches()
        prog["active"] = False


def _start_snapshot(launch_id, cfg, keywords, domains):
    with LAUNCH_LOCK:
        SNAP_PROGRESS[launch_id] = {"active": True, "done": 0,
                                    "total": len(keywords), "cancel": Event(),
                                    "started": time.time()}
    Thread(target=_launch_snapshot, args=(launch_id, cfg, keywords, domains),
           daemon=True).start()


def _snap_cfg(data):
    """cfg для съёма позиций из тела запроса (переключатель Live/XML + глубина)."""
    user = (data.get("user") or "").strip()
    key = (data.get("key") or "").strip()
    if not user or not key:
        return None, "Укажите User ID и API key из кабинета xmlstock."
    engine = data.get("engine") or "yandex"
    family = "google" if str(engine).startswith("google") else "yandex"
    xml = data.get("snap_engine") == "xml"
    endpoint = ENDPOINTS[family if xml else family + "_live"]

    def clamp(v, lo, hi, d):
        try:
            return max(lo, min(hi, int(v)))
        except (TypeError, ValueError):
            return d

    cfg = {
        "user": user, "key": key, "endpoint": endpoint, "live": not xml,
        "lr": (data.get("lr") or "").strip(),
        "domain": (data.get("domain") or "").strip(),
        "device": (data.get("device") or "").strip(),
        "top_n": clamp(data.get("depth"), 10, 100, 50),
        "workers": clamp(data.get("workers"), 1, 10, 5),
        "delay": 0.0, "timeout": 30, "retries": 2,
        "engine_label": "XML" if xml else "Live",
    }
    return cfg, None


@app.route("/api/launches")
def api_launches():
    with LAUNCH_LOCK:
        kws = list(LAUNCHES["keywords"])
        out = []
        for lid, la in LAUNCHES["launches"].items():
            prog = SNAP_PROGRESS.get(lid, {})
            snaps = la.get("snapshots", [])
            out.append({
                "id": lid, "name": la["name"], "domains_count": len(la["domains"]),
                "snapshots_count": len(snaps), "created_at": la["created_at"],
                "last_at": snaps[-1]["at"] if snaps else None,
                "own_kw": len(la.get("keywords") or []),
                "snapping": bool(prog.get("active")),
                "snap_done": prog.get("done", 0), "snap_total": prog.get("total", 0),
            })
        out.sort(key=lambda x: x["created_at"])
    return jsonify({"keywords": kws, "launches": out})


@app.route("/api/launches/keywords", methods=["POST"])
def api_launches_keywords():
    data = request.get_json(force=True, silent=True) or {}
    kws = _split_unique(data.get("keywords"), 5000)
    with LAUNCH_LOCK:
        LAUNCHES["keywords"] = kws
        _save_launches()
    return jsonify({"keywords": kws})


@app.route("/api/launches/create", methods=["POST"])
def api_launches_create():
    data = request.get_json(force=True, silent=True) or {}
    cfg, err = _snap_cfg(data)
    if err:
        return jsonify({"error": err}), 400
    domains, dseen = [], set()
    for line in (data.get("domains") or "").replace("\r", "\n").split("\n"):
        nd = normalize_domain(line)
        if nd and nd not in dseen:
            dseen.add(nd)
            domains.append(nd)
    if not domains:
        return jsonify({"error": "Добавьте домены запуска (по одному в строке)."}), 400
    own_kw = _split_unique(data.get("launch_keywords"), 5000)   # свои ключи запуска (опц.)
    with LAUNCH_LOCK:
        if data.get("keywords") is not None:
            LAUNCHES["keywords"] = _split_unique(data.get("keywords"), 5000)
            _save_launches()
        keywords = own_kw or list(LAUNCHES["keywords"])
        n = len(LAUNCHES["launches"]) + 1
    if not keywords:
        return jsonify({"error": "Задайте ключи — общие вверху вкладки или свои для запуска."}), 400
    name = (data.get("name") or "").strip() or f"Запуск {n}"
    lid = uuid.uuid4().hex
    with LAUNCH_LOCK:
        LAUNCHES["launches"][lid] = {"id": lid, "name": name,
                                     "created_at": time.time(),
                                     "domains": domains, "keywords": own_kw, "snapshots": []}
        _save_launches()
    _start_snapshot(lid, cfg, keywords, domains)
    return jsonify({"id": lid, "name": name, "domains": len(domains),
                    "keywords": len(keywords), "own_kw": len(own_kw)})


@app.route("/api/launches/<lid>/snapshot", methods=["POST"])
def api_launches_snapshot(lid):
    data = request.get_json(force=True, silent=True) or {}
    cfg, err = _snap_cfg(data)
    if err:
        return jsonify({"error": err}), 400
    with LAUNCH_LOCK:
        la = LAUNCHES["launches"].get(lid)
        if not la:
            return jsonify({"error": "Запуск не найден"}), 404
        if data.get("keywords") is not None:
            LAUNCHES["keywords"] = _split_unique(data.get("keywords"), 5000)
            _save_launches()
        domains = list(la["domains"])
        keywords = la.get("keywords") or list(LAUNCHES["keywords"])
        busy = bool(SNAP_PROGRESS.get(lid, {}).get("active"))
    if busy:
        return jsonify({"error": "Съём по этому запуску уже идёт"}), 409
    if not keywords:
        return jsonify({"error": "Сначала задайте ключи."}), 400
    _start_snapshot(lid, cfg, keywords, domains)
    return jsonify({"ok": True})


@app.route("/api/launches/<lid>/delete", methods=["POST"])
def api_launches_delete(lid):
    with LAUNCH_LOCK:
        LAUNCHES["launches"].pop(lid, None)
        SNAP_PROGRESS.pop(lid, None)
        _save_launches()
    return jsonify({"ok": True})


@app.route("/api/launches/<lid>")
def api_launch_detail(lid):
    with LAUNCH_LOCK:
        la = LAUNCHES["launches"].get(lid)
        if not la:
            return jsonify({"error": "Запуск не найден"}), 404
        rows = launch_report_rows(la, la.get("keywords") or LAUNCHES["keywords"])
        snaps = [{"at": s["at"], "engine": s.get("engine"), "depth": s.get("depth")}
                 for s in la["snapshots"]]
        return jsonify({"id": lid, "name": la["name"], "domains": la["domains"],
                        "snapshots": snaps, "rows": rows})


@app.route("/api/launches/export")
def api_launches_export():
    try:
        from openpyxl import Workbook
    except Exception:
        return jsonify({"error": "openpyxl не установлен"}), 400
    with LAUNCH_LOCK:
        launches = sorted([json.loads(json.dumps(la)) for la in LAUNCHES["launches"].values()],
                          key=lambda x: x["created_at"])
        gkw = list(LAUNCHES["keywords"])
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Сводка")
    ws.append(["Запуск", "Доменов", "Ключей", "Съёмов", "Создан", "Последний съём"])
    for la in launches:
        snaps = la["snapshots"]
        ekw = la.get("keywords") or gkw
        ws.append([la["name"], len(la["domains"]), len(set(ekw) | set(_launch_keywords(la))),
                   len(snaps), _fmt_ts(la["created_at"]), _fmt_ts(snaps[-1]["at"]) if snaps else "—"])

    # Лидерборд: домены по всем запускам (по последнему снимку каждого)
    board = {}
    for la in launches:
        snaps = la["snapshots"]
        if not snaps:
            continue
        for kw, dommap in (snaps[-1].get("positions") or {}).items():
            for dom, p in (dommap or {}).items():
                if not isinstance(p, int):
                    continue
                b = board.setdefault(dom, {"hits": 0, "t3": 0, "t10": 0, "best": None, "la": set()})
                b["hits"] += 1
                b["t3"] += 1 if p <= 3 else 0
                b["t10"] += 1 if p <= 10 else 0
                b["best"] = p if b["best"] is None else min(b["best"], p)
                b["la"].add(la["name"])
    wsb = wb.create_sheet("Лидерборд")
    wsb.append(["Домен", "Запуск(и)", "Ключей в топе", "Топ-3", "Топ-10", "Лучшая позиция"])
    for dom, b in sorted(board.items(), key=lambda kv: (-kv[1]["hits"], kv[1]["best"] or 9999)):
        wsb.append([dom, ", ".join(sorted(b["la"])), b["hits"], b["t3"], b["t10"], b["best"]])

    used = set()
    for la in launches:
        wsl = wb.create_sheet(_sheet_name(la["name"], used))
        snaps = la["snapshots"]
        doms = la["domains"]
        rows = launch_report_rows(la, la.get("keywords") or gkw)
        kw_order, look = [], {}
        for r in rows:
            if r["keyword"] not in kw_order:
                kw_order.append(r["keyword"])
            look[(r["keyword"], r["domain"])] = r

        def matrix_block(title, getter):
            wsl.append([title])
            wsl.append(["Ключ \\ Домен"] + list(doms))
            for kw in kw_order:
                line = [kw]
                for dm in doms:
                    r = look.get((kw, dm))
                    v = getter(r) if r else None
                    line.append(v if v is not None else "")
                wsl.append(line)
            wsl.append([])

        def snapshot_block(title, si):
            # по каждому домену две колонки рядом: позиция и URL страницы (текстом)
            wsl.append([title])
            hdr = ["Ключ"]
            for dm in doms:
                hdr += [f"{dm} — поз", f"{dm} — URL"]
            wsl.append(hdr)
            for kw in kw_order:
                line = [kw]
                for dm in doms:
                    r = look.get((kw, dm))
                    pos = r["positions"][si] if r else None
                    url = (r.get("urls") or [None] * len(snaps))[si] if r else None
                    line.append(pos if pos is not None else "")
                    line.append(url or "")
                wsl.append(line)
            wsl.append([])

        if snaps:
            # Сначала — позиции (и URL рядом) по каждому снимку
            for si, s in enumerate(snaps):
                tag = f"{_fmt_ts(s['at'])} {s.get('engine') or ''}".strip()
                snapshot_block(f"Снимок {tag} (позиция + URL страницы рядом)", si)
            matrix_block("Среднее по съёмам — позиции", lambda r: r["avg"])
            # В конце — смена URL по съёмам (только URL, чтобы отследить ротацию)
            if len(snaps) >= 2:
                wsl.append(["Смена URL по съёмам (какой страницей домен ранжируется)"])
                wsl.append(["Ключ", "Домен"]
                           + [f"URL · {_fmt_ts(s['at'])}" for s in snaps] + ["менялся?"])
                for r in rows:
                    useq = r.get("urls") or []
                    present = [u for u in useq if u]
                    if not present:
                        continue
                    changed = "да" if len(set(present)) > 1 else "нет"
                    wsl.append([r["keyword"], r["domain"]]
                               + [u or "" for u in useq] + [changed])
                wsl.append([])
        else:
            wsl.append(["Съёмов нет"])
    if not launches:
        wb.create_sheet("Пусто").append(["Пока нет запусков"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(buf, as_attachment=True, download_name=f"launches_{stamp}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _open_browser(url):
    try:
        webbrowser.open(url)
    except Exception:
        pass


if __name__ == "__main__":
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "5000"))
    url = f"http://{host if host != '0.0.0.0' else '127.0.0.1'}:{port}"
    print(f"\n  XMLStock SERP запущен → {url}\n  (Ctrl+C для остановки)\n")
    if os.environ.get("NO_BROWSER") != "1":
        Thread(target=lambda: (time.sleep(1.0), _open_browser(url)), daemon=True).start()
    app.run(host=host, port=port, threaded=True, debug=False)
