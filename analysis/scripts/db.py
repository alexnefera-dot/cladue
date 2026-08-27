import openpyxl,io,glob,collections,json,core,cfg
KW=core.KW; EXCL=core.EXCL_BRAND
def tier(v): return "ВЧ" if v>=1_000_000 else ("СЧ" if v>=700_000 else "НЧ")
best={}
for p in glob.glob('launches*.xlsx'):
    wb=openpyxl.load_workbook(io.BytesIO(open(p,'rb').read()),read_only=True,data_only=True)
    for sn in wb.sheetnames:
        ws=wb[sn]
        if ws.max_row<500: continue
        if sn not in best or ws.max_row>best[sn][0]: best[sn]=(ws.max_row,p)
    wb.close()
byfile=collections.defaultdict(list)
for sn,(r,p) in best.items(): byfile[p].append(sn)
DOM={}   # domain -> {sheet, lab, brands:{b:{tier,vol,keys:[(p,q)]}}}
for p,sns in byfile.items():
    wb=openpyxl.load_workbook(io.BytesIO(open(p,'rb').read()),data_only=True)
    for sn in sns:
        rows=list(wb[sn].iter_rows(values_only=True))
        allh=[i for i,r in enumerate(rows) if r[0] and isinstance(r[0],str) and r[0].strip().startswith('Ключ \\')]
        snh=[h for h in allh if 'Снимок' in str(rows[h-1][0])]
        if not snh: continue
        hdr=[str(c).strip().lower() for c in rows[snh[0]][1:] if c not in (None,'')]
        cand=[]
        for h in snh:
            nxt=[i for i in allh if i>h]; end=(nxt[0]-1) if nxt else len(rows)
            b=[r for r in rows[h+1:end] if isinstance(r[0],str) and r[0].strip().lower()!='ключ']
            q4=max(len(b)//4,1)
            qs=[sum(1 for r in b[i*q4:(i+1)*q4] if any(v is not None for v in r[1:1+len(hdr)])) for i in range(4)]
            cand.append((b,str(rows[h-1][0]).replace('Снимок ','').replace(' XML',''),
                         qs[3]==0 and qs[0]>=5 and sum(qs)>=10))
        ok=[c for c in cand if not c[2]] or cand
        body,lab,_=ok[-1]
        per=collections.defaultdict(lambda: collections.defaultdict(list))
        for r in body:
            q=r[0].strip().lower(); m=KW.get(q)
            if not m: continue
            br,vol,_=m
            if br in EXCL: continue
            for i,d in enumerate(hdr):
                try: pp=int(r[1+i])
                except: continue
                if 1<=pp<=100: per[d][br].append((pp,q,vol))
        for d in hdr:
            bs={}
            for br,ks in per[d].items():
                ks.sort()
                bs[br]={'tier':tier(ks[0][2]),'vol':int(ks[0][2]),
                        't3':sum(1 for x in ks if x[0]<=3),'t10':sum(1 for x in ks if x[0]<=10),
                        't30':sum(1 for x in ks if x[0]<=30),'t100':len(ks),'best':ks[0][0],
                        'keys':[[x[0],x[1]] for x in ks]}
            c=cfg.C.get(sn,{})
            DOM[d]={'sheet':sn,'group':c.get('name',sn),'src':c.get('src','?'),'pages':c.get('pages','?'),
                    'dates':c.get('dates','—'),'img':c.get('img','—'),'acc':c.get('acc','—'),
                    'ids':c.get('ids','—'),'made':c.get('made','—'),'coh':c.get('coh','архив'),
                    'test':c.get('test','—'),'arm':c.get('arm','—'),'lab':lab,'zone':'.'+d.split('.')[-1],
                    'b':bs}
    wb.close()
json.dump(DOM,open('db.json','w'),ensure_ascii=False)
tot=len(DOM); wb_=sum(1 for d in DOM.values() if d['b'])
print("доменов:",tot,"с хотя бы одним брендом в Т100:",wb_)
print("пар домен-бренд:",sum(len(d['b']) for d in DOM.values()))
print("брендов:",len({b for d in DOM.values() for b in d['b']}))
