import openpyxl,io,glob,os,collections,core,json,cfg
KW=core.KW
TARGET={'leebet':'ЛИБЕТ','banda':'БАНДА','trix':'ТРИКС'}
def tier(v): return "ВЧ" if v>=1_000_000 else ("СЧ" if v>=700_000 else "НЧ")
# сколько ключей ядра у каждого целевого бренда
cnt=collections.Counter(); vol={}
apexkeys=[]
for q,(b,v,_) in KW.items():
    cnt[b]+=1; vol[b]=max(vol.get(b,0),v)
    if 'apex' in q or 'апекс' in q: apexkeys.append((q,b))
for b in TARGET: print(f"{b}: ключей в ядре {cnt[b]}, макс объём {vol[b]:,} ({tier(vol[b])})".replace(',',' '))
print("ключей со словом apex/апекс в ядре:",len(apexkeys), apexkeys[:5])
best={}
for p in glob.glob('launches*.xlsx'):
    wb=openpyxl.load_workbook(io.BytesIO(open(p,'rb').read()),read_only=True,data_only=True)
    for sn in wb.sheetnames:
        ws=wb[sn]
        if ws.max_row<500: continue
        if sn not in best or ws.max_row>best[sn][0]: best[sn]=(ws.max_row,p)
    wb.close()
byfile=collections.defaultdict(list)
for sn,(r,p) in best.items(): byfile[p].append(sn)
res=collections.defaultdict(list)   # brand -> [(pos,dom,sheet,lab,q,tier)]
for p,sns in byfile.items():
    wb=openpyxl.load_workbook(io.BytesIO(open(p,'rb').read()),data_only=True)
    for sn in sns:
        rows=list(wb[sn].iter_rows(values_only=True))
        allh=[i for i,r in enumerate(rows) if r[0] and isinstance(r[0],str) and r[0].strip().startswith('Ключ \\')]
        snh=[h for h in allh if 'Снимок' in str(rows[h-1][0])]
        if not snh: continue
        hdr=[str(c).strip().lower() for c in rows[snh[0]][1:] if c not in (None,'')]
        # последний НЕобрезанный снимок
        cand=[]
        for h in snh:
            nxt=[i for i in allh if i>h]; end=(nxt[0]-1) if nxt else len(rows)
            b=[r for r in rows[h+1:end] if isinstance(r[0],str) and r[0].strip().lower()!='ключ']
            q4=max(len(b)//4,1)
            quart=[sum(1 for r in b[i*q4:(i+1)*q4] if any(v is not None for v in r[1:1+len(hdr)])) for i in range(4)]
            trunc=quart[3]==0 and quart[0]>=5 and sum(quart)>=10
            cand.append((h,b,str(rows[h-1][0]).replace('Снимок ','').replace(' XML',''),trunc))
        ok=[c for c in cand if not c[3]] or cand
        h,body,lab,_=ok[-1]
        for r in body:
            q=r[0].strip().lower(); mm=KW.get(q)
            if not mm: continue
            br,v,_=mm
            if br not in TARGET: continue
            for i,d in enumerate(hdr):
                try: pp=int(r[1+i])
                except: continue
                if 1<=pp<=100: res[br].append((pp,d,sn,lab,q,tier(v)))
json.dump({b:sorted(v) for b,v in res.items()},open('brand4.json','w'),ensure_ascii=False)
for b,label in TARGET.items():
    v=sorted(res[b])
    print(f"\n===== {label} ({b}) — {cnt[b]} ключей в ядре, {len(v)} попаданий в Т100 =====")
    dom=collections.defaultdict(list)
    for pp,d,sn,lab,q,t in v: dom[(d,sn,lab)].append((pp,q))
    rank=sorted(dom.items(),key=lambda x:(-sum(1 for p,_ in x[1] if p<=10),-sum(1 for p,_ in x[1] if p<=30),min(p for p,_ in x[1])))
    for (d,sn,lab),ks in rank[:8]:
        t10=sum(1 for p,_ in ks if p<=10); t3=sum(1 for p,_ in ks if p<=3); t30=sum(1 for p,_ in ks if p<=30)
        nm=cfg.C.get(sn,{}).get('name',sn)
        print(f"  {d:14s} Т3={t3:2d} Т10={t10:2d} Т30={t30:2d} Т100={len(ks):2d}  [{nm[:40]}] съём {lab}")
        for pp,q in sorted(ks)[:6]: print(f"        {pp:3d}  {q}")
