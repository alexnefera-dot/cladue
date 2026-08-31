# Парсер экспорта с несколькими снимками на листе + блоком смены URL сверху.
# Лист: «Смена URL по съёмам» → далее блоки «Снимок <метка> ...» с таблицей Ключ | дом — поз | дом — URL.
import json,re,collections
from openpyxl import load_workbook
def depth(u):
    if not isinstance(u,str): return None
    p=u.split('://',1)[-1]; p='/'+p.split('/',1)[1] if '/' in p else '/'
    return len(re.findall(r'/ru(?=/|$)',p))
def parse(path):
    wb=load_workbook(path,read_only=True,data_only=True)
    OUT={}; URLCH={}
    for sn in wb.sheetnames:
        if sn in ('Сводка','Лидерборд'): continue
        rows=list(wb[sn].iter_rows(values_only=True))
        # блок смены URL (если есть)
        if rows and rows[0] and str(rows[0][0]).startswith('Смена URL'):
            ch=[]
            for r in rows[2:]:
                if not r or not r[0] or str(r[0]).startswith('Снимок'): break
                ch.append(dict(q=str(r[0]),dom=r[1],u0=r[2],u1=r[3],f=str(r[4]).strip() if r[4] else ''))
            URLCH[sn.strip()]=ch
        starts=[i for i,r in enumerate(rows) if r and r[0] and str(r[0]).startswith('Снимок')]
        snaps=[]
        for si,st in enumerate(starts):
            end=starts[si+1] if si+1<len(starts) else len(rows)
            # блок «Среднее по съёмам» — это НЕ снимок, обрезаем по нему
            for i in range(st,end):
                if rows[i] and rows[i][0] and str(rows[i][0]).startswith('Средн'):
                    end=i; break
            lab=re.sub(r'^Снимок\s*','',str(rows[st][0])).split(' XML')[0].strip()
            h=next((i for i in range(st,end) if rows[i] and str(rows[i][0]).strip()=='Ключ'),None)
            if h is None: continue
            cols={str(c).split(' — ')[0].strip():i for i,c in enumerate(rows[h][1:],1)
                  if c and '— поз' in str(c)}
            per=collections.defaultdict(list)
            for r in rows[h+1:end]:
                if not r or not r[0] or str(r[0]).startswith(('Средн','Снимок')): continue
                q=str(r[0]).strip()
                for dom,ci in cols.items():
                    p=r[ci]
                    if p in (None,''): continue
                    try: p=int(p)
                    except: continue
                    u=r[ci+1] if isinstance(r[ci+1],str) else None
                    per[dom].append(dict(q=q,p=p,u=u,d=depth(u)))
            snaps.append(dict(lab=lab,doms=list(cols),per=dict(per)))
        if snaps: OUT[sn.strip()]=snaps
    return OUT,URLCH
if __name__=='__main__':
    import sys
    OUT,URLCH=parse(sys.argv[1] if len(sys.argv)>1 else 'L01.xlsx')
    json.dump(OUT,open('p01.json','w'),ensure_ascii=False)
    json.dump(URLCH,open('p01_url.json','w'),ensure_ascii=False)
    for sn,snaps in OUT.items():
        print('==',sn)
        for s in snaps:
            t=lambda n: sum(1 for d,ks in s['per'].items() for k in ks if k['p']<=n)
            print(f"   {s['lab']:<14} доменов {len(s['doms']):>2}  Т3={t(3):>3} Т10={t(10):>4} "
                  f"Т30={t(30):>4} Т100={sum(len(v) for v in s['per'].values()):>5}")
