# Парсер плоской выгрузки конверсий (Book1.xlsx, 29-31.08):
# ts | рега/деп | uid | движок | гео | бренд.домен.зона | реферер | UA | IP
import json,collections
from openpyxl import load_workbook
ws=load_workbook('conv3.xlsx',read_only=True,data_only=True)['Sheet1']
EV=[]
for r in ws.iter_rows(values_only=True):
    if not r or not r[0]: continue
    ts=str(r[0]); host=str(r[5] or '')
    parts=host.split('.')
    brand=parts[0] if len(parts)>2 else None
    dom='.'.join(parts[1:]) if len(parts)>2 else host
    EV.append(dict(ts=ts,d=ts[8:10]+'.'+ts[5:7],h=int(ts[11:13]),
        type='dep' if str(r[1]).startswith('деп') else 'reg',
        uid=r[2],eng=r[3],geo=r[4],host=host,brand=brand,dom=dom,ref=r[6],ua=r[7],ip=r[8]))
json.dump(EV,open('conv3.json','w'),ensure_ascii=False)
print('событий',len(EV),collections.Counter(e['type'] for e in EV))
print('дни',sorted(collections.Counter(e['d'] for e in EV).items()))
print('движки',collections.Counter(e['eng'] for e in EV))
