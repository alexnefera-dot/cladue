import openpyxl,io,collections,json,statistics as st,core
KW=core.KW; EXCL=core.EXCL_BRAND
def tier(v): return "ВЧ" if v>=1_000_000 else ("СЧ" if v>=700_000 else "НЧ")
wb=openpyxl.load_workbook(io.BytesIO(open('launches21.xlsx','rb').read()),data_only=True)
OUT={}
for sn in wb.sheetnames:
    ws=wb[sn]
    if ws.max_row<200: continue
    rows=list(ws.iter_rows(values_only=True))
    hdrs=[i for i,r in enumerate(rows) if r[0] and str(r[0]).strip()=='Ключ' and any(c and '— поз' in str(c) for c in r[1:])]
    snaps=[]
    for k,h in enumerate(hdrs):
        lab=None
        for j in range(h-1,max(h-3,-1),-1):
            if rows[j][0] and str(rows[j][0]).startswith('Снимок'):
                lab=str(rows[j][0]).replace('Снимок ','').split(' XML')[0]; break
        cols={}
        for i,c in enumerate(rows[h][1:],start=1):
            if c and '— поз' in str(c): cols[str(c).split(' — ')[0].strip().lower()]=i
        end=hdrs[k+1]-2 if k+1<len(hdrs) else len(rows)
        for j in range(h+1,end):
            r0=rows[j][0]
            if r0 and isinstance(r0,str) and ('Среднее' in r0 or r0.strip()=='Ключ'): end=j; break
        body=[r for r in rows[h+1:end] if isinstance(r[0],str) and r[0].strip() and r[0].strip().lower()!='ключ']
        per=collections.defaultdict(list); urls=collections.defaultdict(collections.Counter)
        for r in body:
            q=r[0].strip().lower(); m=KW.get(q)
            if not m: continue
            br,vol,_=m
            if br in EXCL: continue
            for d,ci in cols.items():
                try: p=int(r[ci])
                except: continue
                if not (1<=p<=100): continue
                u=r[ci+1] if ci+1<len(r) else None
                per[d].append({'p':p,'b':br,'t':tier(vol),'q':q,'v':vol,'u':str(u) if u else None})
                if u: urls[d][str(u)]+=1
        snaps.append({'lab':lab,'nkeys':len(body),'doms':list(cols),'per':{d:per[d] for d in cols},
                      'urls':{d:urls[d].most_common(12) for d in cols}})
    OUT[sn]=snaps
    print(f"{sn!r}: снимков={len(snaps)} ключей={snaps[0]['nkeys']} доменов={len(snaps[0]['doms'])} метки={[s['lab'] for s in snaps]}")
json.dump(OUT,open('p21.json','w'),ensure_ascii=False,default=str)
