# Данные для отчёта по пулам 31.08-01.09: бренды, ключи, вложенность, страницы.
import json,re,csv,collections,statistics as st
D=json.load(open('p01.json'))
TIER={}; VOL={}
with open('/home/user/cladue/analysis/keys/keys_stats.csv',encoding='utf-8-sig') as f:
    for r in csv.DictReader(f,delimiter=';'):
        q=r['ключ'].strip()
        TIER[q]=(r['бренд'],r['тир'])
        try: VOL[r['бренд']]=int(r['частотность'])
        except: pass
def parts(u):
    if not isinstance(u,str): return None,None,None
    rest=u.split('://',1)[-1]; host=rest.split('/',1)[0]
    path='/'+rest.split('/',1)[1] if '/' in rest else '/'
    segs=[s for s in path.split('/') if s]
    pg=[s for s in segs if s!='ru']
    return host,('/'+'/'.join(pg) if pg else '/'),len([s for s in segs if s=='ru'])
POOLS=[]
CFG={'Generator_11page_test .com (ru':dict(name='Generator_11page_test · .com',cap='потолок 20',
        plat='dorgen com',ld='31.08 16:39',ids='1055-1063',pages='11'),
     'Generator_11page_test .net (ru':dict(name='nabor-244…253 · .net',cap='без потолка',
        plat='dorgen.net',ld='31.08 19:30–22:45 (точно не прислано)',ids='nabor-244…253',pages='не прислано'),
     'apex, banda':dict(name='apex, banda (узкое ядро 70 ключей)',cap='потолок 20',
        plat='dorgen com',ld='31.08 16:39',ids='1055-1063',pages='11')}
ROWS=[]
for sn,snaps in D.items():
    cfg=CFG.get(sn,dict(name=sn,cap='?',plat='?',ld='?',ids='?',pages='?'))
    labs=[s['lab'] for s in snaps]
    last=snaps[-1]
    per_dom=[]
    for d in snaps[0]['doms']:
        tl=[]
        for s in snaps:
            ks=s['per'].get(d,[])
            tl.append(dict(lab=s['lab'],t3=sum(1 for k in ks if k['p']<=3),
                t10=sum(1 for k in ks if k['p']<=10),t30=sum(1 for k in ks if k['p']<=30),t100=len(ks)))
        ks=last['per'].get(d,[])
        ds=[k['d'] for k in ks if k['d'] is not None]
        pgs=collections.Counter(); subs=collections.Counter(); brs=collections.Counter()
        for k in ks:
            ho,pg,dp=parts(k['u'])
            if pg: pgs[pg]+=1
            if ho: subs[ho.split('.')[0]]+=1
            b=TIER.get(k['q'],(None,None))[0]
            if b and k['p']<=10: brs[b]+=1
        per_dom.append(dict(d=d,tl=tl,best=min([k['p'] for k in ks],default=None),
            dmax=max(ds) if ds else 0,dmed=round(st.median(ds)) if ds else 0,
            dover=round(100*sum(1 for x in ds if x>20)/len(ds)) if ds else 0,
            nsub=len(subs),pages=pgs.most_common(6),brands=brs.most_common(8)))
        for k in ks:
            ho,pg,dp=parts(k['u'])
            b,t=TIER.get(k['q'],(None,None))
            ROWS.append(dict(pool=cfg['name'],cap=cfg['cap'],dom=d,q=k['q'],p=k['p'],
                b=b,t=t,vol=VOL.get(b,0),sub=(ho.split('.')[0] if ho else None),
                page=pg,d=dp,u=k['u']))
    tot=[]
    for s in snaps:
        f=lambda n: sum(1 for ks in s['per'].values() for k in ks if k['p']<=n)
        tot.append(dict(lab=s['lab'],t3=f(3),t10=f(10),t30=f(30),
                        t100=sum(len(v) for v in s['per'].values())))
    ds=[k['d'] for ks in last['per'].values() for k in ks if k['d'] is not None]
    pgs=collections.Counter(); brs=collections.Counter(); tie=collections.Counter(); tie10=collections.Counter()
    for ks in last['per'].values():
        for k in ks:
            ho,pg,dp=parts(k['u'])
            if pg: pgs[pg]+=1
            b,t=TIER.get(k['q'],(None,None))
            if b: brs[b]+= (1 if k['p']<=10 else 0)
            if t: tie[t]+=1; tie10[t]+= (1 if k['p']<=10 else 0)
    # эффект смены адреса между двумя последними съёмами
    eff=None
    if len(snaps)>=2:
        a,b2=snaps[-2],snaps[-1]; ch=[];sm=[]
        for d in a['doms']:
            A={k['q']:k for k in a['per'].get(d,[])}; B={k['q']:k for k in b2['per'].get(d,[])}
            for q in set(A)&set(B):
                if A[q]['d'] is None or B[q]['d'] is None: continue
                (ch if A[q]['d']!=B[q]['d'] else sm).append(B[q]['p']-A[q]['p'])
        eff=dict(ch=len(ch),chmed=round(st.median(ch)) if ch else None,
                 sm=len(sm),smmed=round(st.median(sm)) if sm else None,
                 keep=sum(len({k['q'] for k in a['per'].get(d,[])}&{k['q'] for k in b2['per'].get(d,[])}) for d in a['doms']),
                 was=sum(len(a['per'].get(d,[])) for d in a['doms']))
    POOLS.append(dict(**cfg,sheet=sn,n=len(snaps[0]['doms']),labs=labs,tot=tot,dom=per_dom,
        dmax=max(ds) if ds else 0,dmed=round(st.median(ds)) if ds else 0,
        dover=round(100*sum(1 for x in ds if x>20)/len(ds)) if ds else 0,
        dhist=[[lo,sum(1 for x in ds if lo<=x<=hi)] for lo,hi in [(0,0),(1,5),(6,10),(11,20),(21,30),(31,40),(41,99)]],
        pgtop=pgs.most_common(10),brands=brs.most_common(25),tier=dict(tie),tier10=dict(tie10),eff=eff))
json.dump(dict(pools=POOLS,rows=ROWS),open('pools_rep.json','w'),ensure_ascii=False)
print('пулов',len(POOLS),'строк ключ×домен',len(ROWS))
for p in POOLS: print(' ',p['name'],p['n'],'дом |',[t['t10'] for t in p['tot']],'| /ru мед',p['dmed'],'макс',p['dmax'],'| страниц',len(p['pgtop']))

# --- процент захода: доля доменов с позициями и доля занятого ядра
CORE={'Generator_11page_test .com (ru':1049,'Generator_11page_test .net (ru':1049,'apex, banda':70}
ENTRY=[]
for sn,snaps in D.items():
    s=snaps[-1]; n=len(s['doms']); core=CORE.get(sn,1049)
    bands=[]
    for lim,lab in [(3,'Т3'),(10,'Т10'),(30,'Т30'),(100,'сотня')]:
        dm=sum(1 for d in s['doms'] if any(k['p']<=lim for k in s['per'].get(d,[])))
        ks={k['q'] for d in s['doms'] for k in s['per'].get(d,[]) if k['p']<=lim}
        bands.append(dict(lab=lab,dm=dm,dsh=round(100*dm/n),keys=len(ks),ksh=round(100*len(ks)/core,1)))
    doms=[]
    for d in sorted(s['doms'],key=lambda x:-len(s['per'].get(x,[]))):
        ks=s['per'].get(d,[])
        doms.append(dict(d=d,t100=len(ks),sh100=round(100*len(ks)/core,1),
            t10=sum(1 for k in ks if k['p']<=10),
            sh10=round(100*sum(1 for k in ks if k['p']<=10)/core,1)))
    ENTRY.append(dict(pool=CFG.get(sn,{}).get('name',sn),cap=CFG.get(sn,{}).get('cap','?'),
        n=n,core=core,bands=bands,doms=doms))
# исторический фон: доля доменов с позициями по всем группам архива
HIST=[]
try:
    h=json.load(open('hist.json'))
    tn=ta=tc=0
    for sheet,v in h.items():
        dom=v['snaps'][-1]['dom']; n=len(dom)
        a=sum(1 for x in dom.values() if x['t10']>0); c=sum(1 for x in dom.values() if x['t100']>0)
        HIST.append(dict(g=sheet.strip(),n=n,t10=a,sh10=round(100*a/n),t100=c,sh100=round(100*c/n)))
        tn+=n; ta+=a; tc+=c
    HIST.sort(key=lambda x:-x['sh10'])
    HTOT=dict(n=tn,t10=ta,sh10=round(100*ta/tn),t100=tc,sh100=round(100*tc/tn))
except FileNotFoundError:
    HTOT=None
J=json.load(open('pools_rep.json'))
J.update(entry=ENTRY,hist=HIST,htot=HTOT)
json.dump(J,open('pools_rep.json','w'),ensure_ascii=False)
print('заход:',[(e['pool'],[(b['lab'],b['dsh'],b['ksh']) for b in e['bands']]) for e in ENTRY])
print('архив:',HTOT)

# --- полный список ключей ядра с тиром (все строки листа, даже без позиций)
from openpyxl import load_workbook
import os
CORELIST=[]
for fn in ['L02.xlsx']:
    if not os.path.exists(fn): continue
    wb=load_workbook(fn,read_only=True,data_only=True)
    for sname in wb.sheetnames:
        if sname in ('Сводка','Лидерборд'): continue
        rows=list(wb[sname].iter_rows(values_only=True))
        st=next((i for i,r in enumerate(rows) if r and r[0] and str(r[0]).startswith('Снимок')),None)
        if st is None: continue
        h=next(i for i in range(st,len(rows)) if rows[i] and str(rows[i][0]).strip()=='Ключ')
        ks=[]
        for r in rows[h+1:]:
            if not r or not r[0] or str(r[0]).startswith(('Средн','Снимок')): break
            ks.append(str(r[0]).strip())
        CORELIST.append((sname.strip(),ks))
        break
CK=CORELIST[0][1] if CORELIST else []
TC=collections.Counter(TIER.get(q,(None,'—'))[1] for q in CK)
print('ядро:',len(CK),'по тирам:',dict(TC))
# захват ВЧ/СЧ ядра в ТОП-15
VS=[q for q in CK if TIER.get(q,(None,''))[1] in ('ВЧ','СЧ')]
TOPN=15
HIGH=[]
for sn,snaps in D.items():
    if sn=='apex, banda': continue
    s=snaps[-1]
    got={k['q'] for d in s['doms'] for k in s['per'].get(d,[]) if k['p']<=TOPN}
    gv=[q for q in VS if q in got]
    HIGH.append(dict(pool=CFG.get(sn,{}).get('name',sn),cap=CFG.get(sn,{}).get('cap','?'),
        vs=len(VS),got=len(gv),sh=round(100*len(gv)/len(VS),1),
        allk=len(got),allsh=round(100*len(got)/len(CK),1)))
J=json.load(open('pools_rep.json'))
J.update(core=dict(n=len(CK),tiers=dict(TC),vs=len(VS)),high=HIGH,topn=TOPN)
json.dump(J,open('pools_rep.json','w'),ensure_ascii=False)
print('ВЧ+СЧ в ядре:',len(VS)); print(HIGH)

# --- матрица «полоса позиций × частотность»
BANDS=[3,10,15,30,100]
MAT=[]
for sn,snaps in D.items():
    if sn=='apex, banda': continue
    s=snaps[-1]; row=dict(pool=CFG.get(sn,{}).get('name',sn),cap=CFG.get(sn,{}).get('cap','?'),cells={})
    for lim in BANDS:
        got={k['q'] for d in s['doms'] for k in s['per'].get(d,[]) if k['p']<=lim}
        c=collections.Counter(TIER.get(q,(None,'—'))[1] for q in got)
        row['cells'][str(lim)]=dict(vch=c['ВЧ'],sch=c['СЧ'],nch=c['НЧ'],all=len(got))
    MAT.append(row)
J=json.load(open('pools_rep.json'))
J.update(mat=MAT,bands=BANDS,tiersz=dict(TC))
json.dump(J,open('pools_rep.json','w'),ensure_ascii=False)
for m in MAT: print(m['pool'],{k:(v['vch'],v['sch'],v['nch']) for k,v in m['cells'].items()})
