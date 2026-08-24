import openpyxl,io,collections,json,core
KW=core.KW
src=open('full.py').read(); ns={}; exec(src[:src.index('NAME={')],ns); blocks=ns['blocks']
wb=openpyxl.load_workbook(io.BytesIO(open('launches10.xlsx','rb').read()),data_only=True)
best=collections.defaultdict(lambda: 999)   # (dom,brand) -> лучшая позиция за всё наблюдение
for sn in wb.sheetnames:
    if sn=='Сводка': continue
    for b in blocks(wb[sn]):
        for d,dd in b['data'].items():
            for q,p in dd.items():
                m=KW.get(q)
                if not m: continue
                br=m[0]
                k=(d,br)
                if p<best[k]: best[k]=p
cw=openpyxl.load_workbook(io.BytesIO(open('conv.xlsx','rb').read()),data_only=True)
rows=[r for r in cw['Sheet1'].iter_rows(values_only=True) if r[0]]
D=json.load(open('flat.json',encoding='utf-8'))
KNOWN={d['d'] for d in D['doms']}
GN={d['d']:d['gname'] for d in D['doms']}
buckets=collections.Counter(); det=[]
for r in rows:
    h=str(r[5]).strip().lower(); p=h.split('.')
    if h in ('yandex.ru','ru.search.yahoo.com','alice.yandex.ru'): continue
    dom='.'.join(p[-2:]); br='.'.join(p[:-2])
    if dom not in KNOWN: continue
    bp=best[(dom,br)]
    lvl='ТОП-3' if bp<=3 else 'ТОП-10' if bp<=10 else 'ТОП-30' if bp<=30 else 'ТОП-100' if bp<=100 else 'нет в ядре/не отслеж.'
    buckets[lvl]+=1
    det.append((dom,br,r[1],bp,lvl,GN[dom]))
print('Конверсии по лучшей позиции бренда на этом домене за всё наблюдение (все замеры, весь диапазон Т1-100):')
for l in ['ТОП-3','ТОП-10','ТОП-30','ТОП-100','нет в ядре/не отслеж.']:
    print('   %-24s %3d'%(l,buckets[l]))
print()
print('--- те, где бренда нет в ядре вообще ---')
nk=collections.Counter(b for d,b,k,bp,l,g in det if l.startswith('нет'))
print(' ',', '.join('%s×%d'%(b,n) for b,n in nk.most_common()))
inkw=set(v[0] for v in KW.values())
print()
print('  из них бренд отсутствует в ядре запросов:',sorted({b for d,b,k,bp,l,g in det if l.startswith('нет') and b not in inkw}))
print('  бренд есть в ядре, но домен по нему не ранжировался:',sorted({b for d,b,k,bp,l,g in det if l.startswith('нет') and b in inkw}))
