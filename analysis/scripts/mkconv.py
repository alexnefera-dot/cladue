import openpyxl,io,collections,json,datetime,core
KW=core.KW
src=open('full.py').read(); ns={}; exec(src[:src.index('NAME={')],ns); blocks=ns['blocks']
D=json.load(open('flat.json',encoding='utf-8'))
GN={d['d']:d['gname'] for d in D['doms']}
PG={d['gname']:d['pages'] for d in D['doms']}
ZN={d['d']:d['zone'] for d in D['doms']}
# лучшая позиция бренда на домене за всё наблюдение
wb=openpyxl.load_workbook(io.BytesIO(open('launches10.xlsx','rb').read()),data_only=True)
best=collections.defaultdict(lambda:999)
for sn in wb.sheetnames:
    if sn=='Сводка': continue
    for b in blocks(wb[sn]):
        for d,dd in b['data'].items():
            for q,p in dd.items():
                m=KW.get(q)
                if m and p<best[(d,m[0])]: best[(d,m[0])]=p
cw=openpyxl.load_workbook(io.BytesIO(open('conv.xlsx','rb').read()),data_only=True)
rows=[r for r in cw['Sheet1'].iter_rows(values_only=True) if r[0]]
SE={'yandex.ru','ru.search.yahoo.com','alice.yandex.ru'}
ev=[]
for r in rows:
    h=str(r[5]).strip().lower()
    if h in SE: continue
    p=h.split('.')
    ev.append({'t':r[0],'k':r[1],'dom':'.'.join(p[-2:]),'br':'.'.join(p[:-2]),'c':str(r[4]),'cid':r[3]})
LAUNCH={'Generator_11page':'20.08 01:25','7page_yandex':'20.08 01:25','Generator_11page_2':'20.08 01:25',
 'Generator_v5':'20.08 01:25','Generator_v4_2':'20.08 01:25','generator v4':'20.08 01:25',
 '12pages_withdate · Theme1':'20.08 17:00','12pages_withdate · Theme2':'20.08 17:00',
 '12pages_nodate':'20.08 17:00','7pages_nodate':'20.08 17:00','kostoreznaya1 · имена':'20.08 17:00',
 'nabor28gotovyi · наборы':'20.08 17:00','Generation 50':'20.08 22:00',
 'Generator_11page_21.08 (обе партии)':'22.08 01:18','nabor-53 · наборы':'22.08 01:18',
 'NEW50_5_12pages_withdate':'22.08 01:18','NEW50_5_7pages_withdate':'22.08 01:18',
 'Generator_11page_img':'22.08 23:30','NEW50_5_12pages_nodate':'22.08 23:04','NEW50_5_7pages_nodate':'22.08 23:04'}
END=datetime.datetime(2026,8,24,12,15)
def dt(s):
    d,t=s.split(); dd,mm=d.split('.'); hh,mi=t.split(':')
    return datetime.datetime(2026,int(mm),int(dd),int(hh),int(mi))
gr=collections.defaultdict(lambda: collections.Counter()); gd=collections.defaultdict(set)
dr=collections.defaultdict(lambda: collections.Counter()); dbr=collections.defaultdict(collections.Counter)
for e in ev:
    dr[e['dom']][e['k']]+=1; dbr[e['dom']][e['br']]+=1
    g=GN.get(e['dom'])
    if g: gr[g][e['k']]+=1; gd[g].add(e['dom'])
groups=[]
for g,L in LAUNCH.items():
    n=len([d for d in D['doms'] if d['gname']==g]); days=(END-dt(L)).total_seconds()/86400
    t10=sum(d['t10'] for d in D['doms'] if d['gname']==g); t100=sum(d['t100'] for d in D['doms'] if d['gname']==g)
    c=gr[g]
    groups.append({'g':g,'pages':PG[g],'n':n,'days':round(days,1),'reg':c['reg'],'dep':c['dep'],
      'nd':len(gd[g]),'rpd':c['reg']/n/days,'t10':t10,'t100':t100,
      'r100':100*c['reg']/t10 if t10 else None})
groups.sort(key=lambda x:-x['rpd'])
# --- детализация: по каждой группе все её домены и все события ---
inkw={v[0] for v in KW.values()}
DEV=collections.defaultdict(list)
for e in sorted(ev,key=lambda x:x['t']):
    DEV[e['dom']].append(e)
BR={d['d']:{b['b']:b for b in d['brands']} for d in D['doms']}
def evlist(dom):
    out=[]
    for e in DEV.get(dom,[]):
        bp=best[(dom,e['br'])]
        out.append({'t':e['t'].strftime('%d.%m %H:%M'),'k':e['k'],'br':e['br'],'c':e['c'],
          'p':None if bp>100 else bp,'incore':e['br'] in inkw,
          'n10':BR.get(dom,{}).get(e['br'],{}).get('n',0)})
    return out
for g in groups:
    ds=[]
    for d in D['doms']:
        if d['gname']!=g['g']: continue
        c=dr.get(d['d'],collections.Counter())
        ds.append({'d':d['d'],'zone':d['zone'],'cont':d.get('cont'),'t10':d['t10'],'t30':d['t30'],
          't100':d['t100'],'t3':d['t3'],'vch':d['vch'],'sch':d['sch'],'nb':d['nb'],
          'reg':c['reg'],'dep':c['dep'],'ev':evlist(d['d'])})
    ds.sort(key=lambda x:(-(x['reg']+x['dep']),-x['t10']))
    g['ds']=ds
    bb=collections.Counter()
    for x in ds:
        for e in x['ev']: bb[e['br']]+=1
    g['brands']=bb.most_common()
# домены
dl=[]
for d,c in dr.items():
    known=d in GN
    dl.append({'d':d,'g':GN.get(d),'zone':ZN.get(d,'.'+d.split('.')[-1]),'reg':c['reg'],'dep':c['dep'],
      't10':next((x['t10'] for x in D['doms'] if x['d']==d),None),
      't3':next((x['t3'] for x in D['doms'] if x['d']==d),None),
      'brands':dbr[d].most_common(6),'known':known})
dl.sort(key=lambda x:(-(x['reg']+x['dep']),x['d']))
for d in dl: d['ev']=evlist(d['d'])
# атрибуция по позиции бренда
buck=collections.Counter()
inkw={v[0] for v in KW.values()}
nob=collections.Counter(); nor=collections.Counter()
for e in ev:
    if e['dom'] not in GN: continue
    bp=best[(e['dom'],e['br'])]
    l='ТОП-3' if bp<=3 else 'ТОП-10' if bp<=10 else 'ТОП-30' if bp<=30 else 'ТОП-100' if bp<=100 else ('бренда нет в ядре' if e['br'] not in inkw else 'не ранжировался')
    buck[l]+=1
    if l=='бренда нет в ядре': nob[e['br']]+=1
    if l=='не ранжировался': nor[e['br']]+=1
fmt=collections.defaultdict(lambda: {'reg':0,'dep':0,'dd':0.0,'n':0,'t10':0})
for g,L in LAUNCH.items():
    p=PG[g]; b='12 страниц' if p.startswith('12') else '11 страниц' if p.startswith('11') else '7 страниц' if p.startswith('7') else 'Generation 50 (?)'
    n=len([d for d in D['doms'] if d['gname']==g])
    fmt[b]['dd']+=n*(END-dt(L)).total_seconds()/86400; fmt[b]['n']+=n
    fmt[b]['t10']+=sum(d['t10'] for d in D['doms'] if d['gname']==g)
    fmt[b]['reg']+=gr[g]['reg']; fmt[b]['dep']+=gr[g]['dep']
fm=[{'b':b,'n':a['n'],'dd':round(a['dd'],1),'reg':a['reg'],'dep':a['dep'],'t10':a['t10'],
     'rpd':a['reg']/a['dd'],'r100':100*a['reg']/a['t10'] if a['t10'] else 0} for b,a in fmt.items()]
fm.sort(key=lambda x:-x['rpd'])
brands=collections.defaultdict(lambda: collections.Counter())
for e in ev: brands[e['br']][e['k']]+=1
bl=sorted(({'b':b,'reg':c['reg'],'dep':c['dep']} for b,c in brands.items()),key=lambda x:-(x['reg']+x['dep']))[:24]
zc=collections.defaultdict(lambda: collections.Counter())
for e in ev: zc['.'+e['dom'].split('.')[-1]][e['k']]+=1
new=[e for e in ev if e['dom'] in GN]; old=[e for e in ev if e['dom'] not in GN]
def agg(x): return {'ev':len(x),'reg':sum(1 for e in x if e['k']=='reg'),'dep':sum(1 for e in x if e['k']=='dep'),'doms':len({e['dom'] for e in x})}
D['conv']={'tot':agg(ev),'new':agg(new),'old':agg(old),'skipped':len(rows)-len(ev),
  'groups':groups,'doms':dl,'fmt':fm,'buck':dict(buck),'nob':nob.most_common(14),'nor':nor.most_common(14),'brands':bl,
  'zones':sorted(({'z':z,'reg':c['reg'],'dep':c['dep']} for z,c in zc.items()),key=lambda x:-(x['reg']+x['dep'])),
  'period':'01.08 → 24.08 12:15'}
json.dump(D,open('flat.json','w',encoding='utf-8'),ensure_ascii=False)
print('conv добавлен |',D['conv']['tot'],'| групп',len(groups),'| доменов',len(dl))
print('атрибуция',dict(buck))
