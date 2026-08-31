# Парсер экспорта launches_20260831_211207.xlsx:
# листы-запуски в формате «Ключ | домен — поз | домен — URL», плюс лист смены URL по 1893.team.
import json,re,collections
from openpyxl import load_workbook
wb=load_workbook('L31.xlsx',read_only=True,data_only=True)
def depth(u):
    if not u or not isinstance(u,str): return None
    p=u.split('://',1)[-1]; p='/'+p.split('/',1)[1] if '/' in p else '/'
    return len(re.findall(r'/ru(?=/|$)',p))
OUT={}
for sn in wb.sheetnames:
    if sn in ('Сводка','Лидерборд','1893.team'): continue
    rows=list(wb[sn].iter_rows(values_only=True))
    hdr=next(i for i,r in enumerate(rows) if r and str(r[0]).strip()=='Ключ')
    lab=str(rows[0][0]).replace('Снимок','').strip()
    cols={}
    for i,c in enumerate(rows[hdr][1:],1):
        if c and '— поз' in str(c): cols[str(c).split(' — ')[0].strip()]=i
    per=collections.defaultdict(list)
    for r in rows[hdr+1:]:
        if not r or not r[0] or str(r[0]).startswith('Средн'): continue
        q=str(r[0]).strip()
        for dom,ci in cols.items():
            p=r[ci]
            if p in (None,''): continue
            try: p=int(p)
            except: continue
            u=r[ci+1] if isinstance(r[ci+1],str) else None
            per[dom].append(dict(q=q,p=p,u=u,d=depth(u)))
    OUT[sn.strip()]=dict(lab=lab,doms=list(cols),per=dict(per))
json.dump(OUT,open('p31.json','w'),ensure_ascii=False)
for sn,v in OUT.items():
    print('==',sn,'|',v['lab'])
    for d in v['doms']:
        ks=v['per'].get(d,[])
        c=lambda n: sum(1 for k in ks if k['p']<=n)
        ds=[k['d'] for k in ks if k['d'] is not None]
        print(f"   {d:<12} Т3={c(3):>3} Т10={c(10):>4} Т30={c(30):>4} Т100={len(ks):>4} "
              f"брендов Т10={len({k['q'] for k in ks if k['p']<=10}):>3} "
              f"/ru макс={max(ds) if ds else 0:>3} мед={sorted(ds)[len(ds)//2] if ds else 0}")
