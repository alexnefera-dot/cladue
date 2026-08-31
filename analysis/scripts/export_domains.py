# Выгрузка всех доменов: группа, контент, id, даты, позиции, трафик, конверсии.
# Пишет domains_export.csv (UTF-8 BOM) и domains_export.xlsx.
import json,re,collections,csv
SP='/tmp/claude-0/-home-user-cladue/7a7c5bac-d634-59c6-bc3f-c4e28ea7944c/scratchpad/'
MD='/home/user/cladue/analysis/launches.md'
db=json.load(open(SP+'db.json')); C=json.load(open(SP+'conv2.json'))
TR={d:v for d,v in C['tr'].items() if v['is_dom']}
EV=C['ev']
rd=collections.Counter(e['dom'] for e in EV if e['type']=='reg')
dp=collections.Counter(e['dom'] for e in EV if e['type']=='dep')
P1=['2139.team','2483.team','ogax.team','byai.team','7186.team','4087.team','2084.team','2304.team','7440.team','0302.team']
P2=['bmtq.team','cnwv.team','dprz.team','fkxb.team','glhd.team','hjsf.team','1524.team','1893.team','2367.team','2745.team','4328.team']
OLD=['3596.team','b8rn.team','c5vt.team','d3mw.team','f9kq.team','f9pb.team','h7nd.team','j2t.team','k6m.team','r9v.team']
CFG27={'7page_27.08 · партия 1':dict(group='',src='7page (автор не указан)',pages='7',dates='—',img='—',acc='—',test='—',ids=''),
       '7page_27.08 · партия 2':dict(group='',src='7page (автор не указан)',pages='7',dates='—',img='—',acc='—',test='—',ids=''),
       'Generator_11page_old_27.08':dict(group='',src='наш генератор · old',pages='11',dates='—',img='старый стиль',acc='—',test='—',ids='')}
G27={**{d:('7page_27.08 · партия 1','27.08','27.08 17:33-17:40','да') for d in P1},
     **{d:('7page_27.08 · партия 2','27.08','',  'нет') for d in P2},
     **{d:('Generator_11page_old_27.08','27.08','27.08 18:22','да') for d in OLD}}
# имена контентов из таблиц launches.md
dom_re=re.compile(r'^[a-z0-9][a-z0-9-]{1,14}\.(team|casino|work|live|click|fun|space|lol|store|shop|net|com|online|site|top)$')
id_re=re.compile(r'^(nabor-)?\d{2,4}$'); dt_re=re.compile(r'^\d{2}\.\d{2}(\s+\d{2}:\d{2}(-\d{2}:\d{2})?)?$')
NAMES={}; cur=None
for ln in open(MD):
    st=ln.strip()
    if st.startswith('#'): cur=st.lstrip('#').strip()
    if not st.startswith('|'): continue
    cells=[c.strip().replace('`','').replace('**','') for c in st.strip('|').split('|')]
    d=[c for c in cells if dom_re.match(c)]
    if len(d)!=1: continue
    dom=d[0]; rest=[c for c in cells if c!=dom]
    iid=next((c for c in rest if id_re.match(c)),'')
    dt=next((c for c in rest if dt_re.match(c)),'')
    nm=next((c for c in rest if c not in (iid,dt) and len(c)>2 and not dom_re.match(c) and not c.startswith('.')),'')
    if not nm and iid.startswith('nabor-'): nm=iid
    rec=dict(name=nm,id=iid,made=dt)
    if dom not in NAMES or (not NAMES[dom]['name'] and nm): NAMES[dom]=rec
# позиции: db.json (до launches20) + последний съём 27.08-групп
POS={}
for d,v in db.items():
    if d=='базовый домен': continue
    bs=v['b'].values()
    POS[d]=dict(lab=v['lab'],t3=sum(x['t3'] for x in bs),t10=sum(x['t10'] for x in bs),
                t30=sum(x['t30'] for x in bs),t100=sum(x['t100'] for x in bs),
                nb=sum(1 for x in bs if x['t10']),
                best=min([x['best'] for x in bs if x['best']],default=None))
S={}
for f in ['p21','p22','p23','p24']:
    for sheet,snaps in json.load(open(SP+f+'.json')).items():
        for s in snaps: S.setdefault((sheet,s['lab']),s)
for sh,lab in [('7page_27.08','28.08 12:43'),('Generator_11page_old_27.08','28.08 12:43')]:
    s=S.get((sh,lab))
    if not s: continue
    for dom,ks in s['per'].items():
        c=lambda n: sum(1 for k in ks if k['p']<=n)
        POS[dom]=dict(lab=lab,t3=c(3),t10=c(10),t30=c(30),t100=len(ks),
                      nb=len({k['b'] for k in ks if k['p']<=10}),
                      best=min([k['p'] for k in ks],default=None))
OVLAUNCH={'NEW50_5_7pages_nodate_21.08 _7…_17':'25.08',
          'КОНТРОЛЬ NEW50_5_7pages_nodate_21.08 _18…_24':'26.08'}
def era(ld):
    if not ld: return 'до 19.08 (старая база)'
    return '19–20.08' if ld in ('19.08','20.08') else '21–27.08'
ROWS=[]
allrows=set(db)-{'базовый домен'}|set(G27)|set(TR)
for d in sorted(allrows):
    v=db.get(d); g27=G27.get(d)
    if g27:
        grp,ld,made,cap=g27
        v=v or CFG27.get(grp)
    elif v:
        grp=v['group']; made=v['made']; ld=OVLAUNCH.get(grp,made[:5]); cap='нет'
    else: grp='старые запуски (до 19.08)'; made=''; ld=''; cap='нет'
    n=NAMES.get(d,{}); p=POS.get(d,{}); t=TR.get(d,{})
    ROWS.append([d,'.'+d.split('.')[-1],era(ld),ld,grp,n.get('name',''),n.get('id','') or (v['ids'] if v else ''),
                 n.get('made','') or made,cap,
                 v['src'] if v else '',v['pages'] if v else '',v['dates'] if v else '',v['img'] if v else '',
                 v['acc'] if v else '',v['test'] if v else '',
                 p.get('lab',''),p.get('t3',''),p.get('t10',''),p.get('t30',''),p.get('t100',''),
                 p.get('nb',''),p.get('best',''),
                 t.get('sub',''),t.get('hits',''),t.get('uniq',''),rd.get(d,0),dp.get(d,0)])
HDR=['домен','зона','эпоха','дата запуска','группа','название контента','id','создан','потолок /ru',
     'источник контента','страниц','даты в тексте','картинки/стиль','аккаунты','тест',
     'последний съём','Т3','Т10','Т30','Т100','брендов в Т10','лучшая позиция',
     'страниц-поддоменов','заходов','посетителей','регистраций','депозитов']
with open(SP+'domains_export.csv','w',encoding='utf-8-sig',newline='') as f:
    w=csv.writer(f,delimiter=';'); w.writerow(HDR); w.writerows(ROWS)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font,Alignment,PatternFill
    from openpyxl.utils import get_column_letter
    wb=Workbook(); ws=wb.active; ws.title='Домены'
    ws.append(HDR)
    for r in ROWS: ws.append(r)
    hf=Font(bold=True,color='FFFFFF'); fill=PatternFill('solid',fgColor='2D3240')
    for c in ws[1]: c.font=hf; c.fill=fill; c.alignment=Alignment(vertical='top',wrap_text=True)
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    widths=[14,8,20,12,34,32,12,17,11,24,9,13,17,10,10,14,6,6,6,7,9,9,10,9,12,12,11]
    for i,w_ in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w_
    wb.save(SP+'domains_export.xlsx')
    print('xlsx ok')
except Exception as e:
    print('xlsx пропущен:',e)
print('строк',len(ROWS))
import collections as cc
print(cc.Counter(r[2] for r in ROWS))
print('с названием контента:',sum(1 for r in ROWS if r[5]))
