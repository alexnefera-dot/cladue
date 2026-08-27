import openpyxl,io,glob,os,json,collections,core
KW=core.KW; EXCL=core.EXCL_BRAND
def tier(v): return "ВЧ" if v>=1_000_000 else ("СЧ" if v>=700_000 else "НЧ")
fs=glob.glob('launches*.xlsx')
best={}
for p in fs:
    wb=openpyxl.load_workbook(io.BytesIO(open(p,'rb').read()),read_only=True,data_only=True)
    for sn in wb.sheetnames:
        ws=wb[sn]
        if ws.max_row<500: continue
        if sn not in best or ws.max_row>best[sn][0]: best[sn]=(ws.max_row,p)
    wb.close()
byfile=collections.defaultdict(list)
for sn,(r,p) in best.items(): byfile[p].append(sn)
OUT={}
for p,sns in byfile.items():
    wb=openpyxl.load_workbook(io.BytesIO(open(p,'rb').read()),data_only=True)
    for sn in sns:
        rows=list(wb[sn].iter_rows(values_only=True))
        allh=[i for i,r in enumerate(rows) if r[0] and isinstance(r[0],str) and r[0].strip().startswith('Ключ \\')]
        snh=[h for h in allh if 'Снимок' in str(rows[h-1][0])]
        if not snh: continue
        hdr=[str(c).strip().lower() for c in rows[snh[0]][1:] if c not in (None,'')]
        snaps=[]
        for si,h in enumerate(snh):
            nxt=[i for i in allh if i>h]; end=(nxt[0]-1) if nxt else len(rows)
            b=[r for r in rows[h+1:end] if isinstance(r[0],str) and r[0].strip().lower()!='ключ']
            q=max(len(b)//4,1)
            quart=[sum(1 for r in b[i*q:(i+1)*q] if any(v is not None for v in r[1:1+len(hdr)])) for i in range(4)]
            per=collections.defaultdict(list)
            for r in b:
                qq=r[0].strip().lower(); mm=KW.get(qq)
                if not mm: continue
                br,vol,_=mm
                if br in EXCL: continue
                for i,d in enumerate(hdr):
                    try: pp=int(r[1+i])
                    except: continue
                    if 1<=pp<=100: per[d].append((pp,br,vol,qq))
            last=(si==len(snh)-1)
            dm={}
            for d in hdr:
                ks=per[d]
                rec={'t3':sum(1 for x in ks if x[0]<=3),'t10':sum(1 for x in ks if x[0]<=10),
                     't30':sum(1 for x in ks if x[0]<=30),'t100':len(ks),
                     'vch':sum(1 for x in ks if x[0]<=10 and tier(x[2])=='ВЧ'),
                     'sch':sum(1 for x in ks if x[0]<=10 and tier(x[2])=='СЧ'),
                     'nb':len({x[1] for x in ks})}
                if last:
                    top=sorted([x for x in ks if x[0]<=10],key=lambda z:(z[0],-z[2]))[:8]
                    rec['keys']=[{'q':x[3],'p':x[0],'b':x[1],'t':tier(x[2])} for x in top]
                dm[d]=rec
            s={'lab':str(rows[h-1][0]).replace('Снимок ','').replace(' XML',''),'quart':quart,'corelen':len(b),'dom':dm}
            if last:
                bc=collections.Counter()
                for d in hdr:
                    if not d.endswith('.team'): continue
                    for x in per[d]:
                        if x[0]<=10: bc[x[1]]+=1
                s['topb']=bc.most_common(12); s['nbrands']=len(bc)
            snaps.append(s)
        OUT[sn]={'file':p,'doms':hdr,'snaps':snaps}
    wb.close()
json.dump(OUT,open('hist.json','w'),ensure_ascii=False)
print("листов:",len(OUT),"размер:",os.path.getsize('hist.json')//1024,"k")
