import openpyxl,io,glob,collections,json,core
KW=core.KW
def tier(v): return "ВЧ" if v>=1_000_000 else ("СЧ" if v>=700_000 else "НЧ")
# все листы, все снимки (в т.ч. обрезанные — они дают только дополнительные свидетельства)
best={}
for p in glob.glob('launches*.xlsx'):
    wb=openpyxl.load_workbook(io.BytesIO(open(p,'rb').read()),read_only=True,data_only=True)
    for sn in wb.sheetnames:
        ws=wb[sn]
        if ws.max_row<500: continue
        if sn not in best or ws.max_row>best[sn][0]: best[sn]=(ws.max_row,p)
    wb.close()
byfile=collections.defaultdict(list)
for sn,(r,p) in best.items(): byfile[p].append(sn)
K=collections.defaultdict(lambda:{'sn':0,'hits':0,'t10':0,'t3':0,'best':999,'doms':set(),'sheets':set(),'seen':0})
nsn=0
for p,sns in byfile.items():
    wb=openpyxl.load_workbook(io.BytesIO(open(p,'rb').read()),data_only=True)
    for sn in sns:
        rows=list(wb[sn].iter_rows(values_only=True))
        allh=[i for i,r in enumerate(rows) if r[0] and isinstance(r[0],str) and r[0].strip().startswith('Ключ \\')]
        snh=[h for h in allh if 'Снимок' in str(rows[h-1][0])]
        if not snh: continue
        hdr=[str(c).strip().lower() for c in rows[snh[0]][1:] if c not in (None,'')]
        for h in snh:
            nsn+=1
            nxt=[i for i in allh if i>h]; end=(nxt[0]-1) if nxt else len(rows)
            for r in rows[h+1:end]:
                if not isinstance(r[0],str): continue
                q=r[0].strip().lower()
                if q=='ключ' or not q: continue
                e=K[q]; e['seen']+=1
                got=False
                for i,d in enumerate(hdr):
                    try: pp=int(r[1+i])
                    except: continue
                    if not (1<=pp<=100): continue
                    got=True; e['hits']+=1; e['doms'].add(d); e['sheets'].add(sn)
                    if pp<e['best']: e['best']=pp
                    if pp<=10: e['t10']+=1
                    if pp<=3: e['t3']+=1
                if got: e['sn']+=1
    wb.close()
out={}
for q,e in K.items():
    m=KW.get(q); br,vol=(m[0],m[1]) if m else (None,0)
    out[q]={'sn':e['sn'],'hits':e['hits'],'t10':e['t10'],'t3':e['t3'],
            'best':(e['best'] if e['best']<999 else None),'nd':len(e['doms']),
            'ns':len(e['sheets']),'seen':e['seen'],'b':br,'v':int(vol) if vol else 0,
            't':tier(vol) if vol else None}
json.dump(out,open('keys.json','w'),ensure_ascii=False)
tot=len(out); rank=sum(1 for v in out.values() if v['hits'])
print(f"снимков просмотрено: {nsn}")
print(f"уникальных ключей во всех выгрузках: {tot}")
print(f"когда-либо ранжировались (1-100): {rank}  ({100*rank/tot:.1f}%)")
print(f"ни разу не показали позицию: {tot-rank}  ({100*(tot-rank)/tot:.1f}%)")
print(f"  из них попадали в ТОП-10: {sum(1 for v in out.values() if v['t10'])}")
print(f"  в ТОП-3: {sum(1 for v in out.values() if v['t3'])}")
nb=sum(1 for v in out.values() if not v['b'])
print(f"ключей без опознанного бренда: {nb}")
